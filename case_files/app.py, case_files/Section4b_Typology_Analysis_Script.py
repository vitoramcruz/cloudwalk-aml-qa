"""
CloudWalk Payments Inc. — AML Transaction Monitoring
Comprehensive Suspect Detection & Ranking Engine

Runs all 11 confirmed typologies against the AMLFT_Analyst_JIM dataset,
produces a composite suspicion score per customer and per transaction,
and outputs:
  - TOP 10 MOST SUSPICIOUS CUSTOMERS (ranked, with narrative + evidence)
  - TOP 30 MOST SUSPICIOUS TRANSACTIONS (ranked)
  - Per-typology detection detail (11 sheets)

Detection logic and default parameters are carried over from
TASK_3_1 (aml_detection_queries.sql / detection_logic_parameter_reference.md)
for consistency across the case file. Where this task's brief added a
calibration instruction not present in Task 3.1 (Typology 8, cash-in/
cash-out), that instruction is implemented and documented inline.

All code and output: English.
Prepared by: AML/BSA Compliance — Transaction Monitoring | CloudWalk AML Case
"""

import json
import re
import warnings
from datetime import timedelta

import numpy as np
import pandas as pd
from rapidfuzz import fuzz

warnings.filterwarnings("ignore")

# ============================================================================
# PORTABILITY CONFIG (as specified in the task brief)
# ============================================================================
# NOTE ON PORTABILITY: point INPUT_FILE at the workbook on your own filesystem
# (e.g. just "AMLFT_Analyst_JIM.xlsx" if it sits next to this script). The
# absolute path below reflects where the source workbook lived when this
# analysis was produced.
INPUT_FILE = "/mnt/project/AMLFT_Analyst_JIM__1_.xlsx"
SHEET_TRANSACTIONS = "Transactions"
SHEET_CUSTOMERS = "Customers_KYC"
SHEET_MERCHANTS = "Merchants_KYB"
SHEET_CARDS = "Cards"
ANALYSIS_DATE = pd.Timestamp("2025-11-07T23:59:59Z")
HIGH_RISK_MCCS = [4829, 5944, 5967, 6011, 6051, 7995]

# ============================================================================
# LOAD DATA
# ============================================================================
print("Loading workbook...")
xls = pd.ExcelFile(INPUT_FILE)
tx = pd.read_excel(xls, SHEET_TRANSACTIONS)
cust = pd.read_excel(xls, SHEET_CUSTOMERS)
merch = pd.read_excel(xls, SHEET_MERCHANTS)
cards = pd.read_excel(xls, SHEET_CARDS)

tx["txn_timestamp_utc"] = pd.to_datetime(tx["txn_timestamp_utc"])
cust["created_at"] = pd.to_datetime(cust["created_at"], errors="coerce")
merch["onboarding_date"] = pd.to_datetime(merch["onboarding_date"], errors="coerce")

DATASET_MAX_TS = tx["txn_timestamp_utc"].max()  # 2025-11-07 23:43:56 UTC

print(f"  Transactions: {len(tx):,} | Customers: {len(cust):,} | "
      f"Merchants: {len(merch):,} | Cards: {len(cards):,}")
print(f"  Date range: {tx['txn_timestamp_utc'].min()} -> {DATASET_MAX_TS}")

# ---- Enrichment: attach merchant + card + customer attributes to every txn
merch_small = merch[[
    "merchant_id", "dba_name", "primary_mcc", "high_risk_mcc_flag",
    "risk_rating", "beneficial_owners_json", "ofac_match_score",
    "incorporation_country",
]].rename(columns={"risk_rating": "merchant_risk_rating"})

cards_small = cards[["card_id", "customer_id", "product", "issue_country"]].rename(
    columns={"customer_id": "card_customer_id"}
)

cust_small = cust[[
    "customer_id", "full_name", "country", "doc_issue_country", "pep_flag",
    "sanctions_match_score", "risk_rating", "expected_avg_ticket_usd",
    "expected_monthly_volume_usd", "kyc_level",
]].rename(columns={"risk_rating": "customer_risk_rating"})

txe = tx.merge(merch_small, on="merchant_id", how="left", validate="m:1")
txe = txe.merge(cust_small, on="customer_id", how="left", validate="m:1")

print("Data loaded and enriched.\n")

# ============================================================================
# COMPOSITE SCORE WEIGHT TABLE
# ============================================================================
# Each typology contributes points to a customer's composite suspicion score
# when the customer is implicated. Weights reflect the inherent severity of
# the typology under BSA/AML risk-tiering (sanctions/self-dealing exposure
# scored highest; single-signal behavioral anomalies scored lower), then are
# scaled within a typology by the magnitude of the evidence (frequency,
# ratio, distinct-party count). This is a transparent, documented rubric —
# not a regulatory formula — used solely to rank triage priority.
TYPOLOGY_WEIGHTS = {
    "T11_FATF_OFAC_JURISDICTION": 32,     # OFAC comprehensive sanctions / FATF black exposure
    "T10_SELF_MERCHANT": 25,              # self-dealing / undisclosed UBO relationship
    "T03_DEVICE_SHARING": 20,             # mule-ring infrastructure signal
    "T09_IP_RING": 18,                    # mule-ring infrastructure signal
    "T08_CASH_IN_CASH_OUT": 16,           # layering (placement -> transit)
    "T01_STRUCTURING": 16,                # intentional evasion pattern
    "T05_PEP_HIGH_RISK_MCC": 13,          # mandatory-EDD population, profile deviation
    "T02_CARD_TESTING": 10,               # fraud-enablement signal
    "T04_GEO_HOPPING_XBORDER": 10,        # cross-border risk-stacking
    "T06_ECOM_NO_3DS": 8,                 # authentication-evasion
    "T07_CHARGEBACK_OUTLIER": 8,          # contextual (merchant-level; customer link)
}

print("Weight table set:", json.dumps(TYPOLOGY_WEIGHTS, indent=2))

# Container for all per-typology detail tables (-> Excel sheets) and for the
# customer-level / transaction-level score contributions used in the final
# ranking.
TYPOLOGY_TABLES = {}
CUSTOMER_HITS = []   # rows: customer_id, typology, weight, magnitude_note, evidence_txn_ids
TXN_HITS = []        # rows: txn_id, typology, weight, note


def sliding_window_group(df, group_col, ts_col, window, min_events, extra_qualify=None,
                          select="earliest", rank_key=None):
    """
    Replicates the SQL 'anchor row -> aggregate forward window' pattern used
    throughout TASK_3_1: for each row, look forward `window` from its
    timestamp within the same group_col, collect all rows in that span, and
    keep groups whose forward-looking window satisfies min_events.

    select="earliest" -> keep the first qualifying window per group
        (matches SQL's `ORDER BY group, window_start` DISTINCT ON, used for
        Query 1 card testing and Query 3 structuring).
    select="best"     -> among ALL qualifying windows for the group, keep the
        one maximizing rank_key(window_df), tie-broken by earliest start
        (matches SQL's `ORDER BY group, metric DESC, window_start` DISTINCT
        ON, used for Query 2 device sharing).
    """
    results = []
    for gval, gdf in df.groupby(group_col):
        gdf = gdf.sort_values(ts_col).reset_index(drop=True)
        ts = gdf[ts_col].values
        n = len(gdf)
        qualifying = []
        for i in range(n):
            end_ts = ts[i] + np.timedelta64(int(window.total_seconds()), "s")
            j = i
            while j < n and ts[j] <= end_ts:
                j += 1
            window_df = gdf.iloc[i:j]
            if extra_qualify is not None and not extra_qualify(window_df):
                continue
            if len(window_df) >= min_events:
                qualifying.append(window_df)
                if select == "earliest":
                    break
        if not qualifying:
            continue
        if select == "earliest":
            results.append((gval, qualifying[0]))
        else:  # "best"
            best_df = max(qualifying, key=rank_key)
            results.append((gval, best_df))
    return results


# ============================================================================
# TYPOLOGY 1 — STRUCTURING
# ============================================================================
# Query 3 params (Task 3.1): min_txns=3, band $980-$995, window_days=7,
# target_mccs=[6051,4829]. Task brief's confirmed finding (104 txns) uses the
# broader "all high-risk MCCs" framing -> validated below and both cuts are
# reported; ranking uses the documented Query-3 target MCCs [6051, 4829]
# (quasi-cash / money-transfer, the channels through which structured value
# is actually converted), consistent with Task 3.1.
print("\n[Typology 1] Structuring...")
STRUCT_MIN_TXNS = 3
STRUCT_LOW, STRUCT_HIGH = 980.00, 995.00
STRUCT_WINDOW = timedelta(days=7)
STRUCT_TARGET_MCCS = [6051, 4829]

band_all_hr = txe[
    (txe["amount_usd"].between(STRUCT_LOW, STRUCT_HIGH))
    & (txe["mcc"].isin(HIGH_RISK_MCCS))
]
band_target = txe[
    (txe["amount_usd"].between(STRUCT_LOW, STRUCT_HIGH))
    & (txe["mcc"].isin(STRUCT_TARGET_MCCS))
]
print(f"  Band $980-995 in ALL high-risk MCCs: {len(band_all_hr)} txns "
      f"(brief's confirmed count: 104)")
print(f"  Band $980-995 in target MCCs [6051,4829] only: {len(band_target)} txns")

struct_windows = sliding_window_group(
    band_all_hr, "customer_id", "txn_timestamp_utc", STRUCT_WINDOW, STRUCT_MIN_TXNS
)
struct_rows = []
for cust_id, wdf in struct_windows:
    struct_rows.append({
        "typology": "T01_STRUCTURING",
        "customer_id": cust_id,
        "window_start": wdf["txn_timestamp_utc"].min(),
        "window_end": wdf["txn_timestamp_utc"].max(),
        "txns_in_window": len(wdf),
        "total_amount_in_window": wdf["amount_usd"].sum(),
        "mccs_in_window": sorted(wdf["mcc"].unique().tolist()),
        "txn_ids": ",".join(wdf["txn_id"].tolist()),
    })
    weight = TYPOLOGY_WEIGHTS["T01_STRUCTURING"] * min(1.5, 1 + 0.1 * (len(wdf) - STRUCT_MIN_TXNS))
    CUSTOMER_HITS.append({
        "customer_id": cust_id, "typology": "T01_STRUCTURING", "weight": weight,
        "evidence": f"{len(wdf)} txns in ${STRUCT_LOW}-${STRUCT_HIGH} band within 7 days "
                    f"(MCCs {sorted(wdf['mcc'].unique().tolist())}); total ${wdf['amount_usd'].sum():,.2f}",
        "txn_ids": wdf["txn_id"].tolist(),
    })
    for _, r in wdf.iterrows():
        TXN_HITS.append({"txn_id": r["txn_id"], "typology": "T01_STRUCTURING",
                          "weight": weight / len(wdf),
                          "note": f"Structuring window ({len(wdf)} txns/7d, MCC {r['mcc']})"})

struct_df = pd.DataFrame(struct_rows).sort_values("txns_in_window", ascending=False)
TYPOLOGY_TABLES["T01_Structuring"] = struct_df
print(f"  Qualifying customers (>=3 txns/7d in high-risk MCC band): {len(struct_df)}")

# ============================================================================
# TYPOLOGY 2 — CARD TESTING
# ============================================================================
# Query 1 params: min_attempts=5, window_minutes=30, threshold_amount=$5.00,
# min_distinct_merchants=3, ECOM channel, at least one approval in-window.
print("\n[Typology 2] Card Testing...")
CT_MIN_ATTEMPTS = 5
CT_WINDOW = timedelta(minutes=30)
CT_THRESHOLD = 5.00
CT_MIN_MERCHANTS = 3

low_value_ecom = txe[(txe["channel"] == "ECOM") & (txe["amount_usd"] < CT_THRESHOLD)]
print(f"  ECOM txns < $5: {len(low_value_ecom)} "
      f"(brief's confirmed universe: 254 txns, 70% decline, 22 unique cards)")

card_windows = sliding_window_group(
    low_value_ecom, "card_id", "txn_timestamp_utc", CT_WINDOW, CT_MIN_ATTEMPTS,
    extra_qualify=lambda w: (w["merchant_id"].nunique() >= CT_MIN_MERCHANTS)
                            and (w["status"] == "approved").any(),
)
ct_rows = []
for card_id, wdf in card_windows:
    cust_id = wdf["customer_id"].iloc[0]
    ct_rows.append({
        "typology": "T02_CARD_TESTING",
        "card_id": card_id,
        "customer_id": cust_id,
        "window_start": wdf["txn_timestamp_utc"].min(),
        "window_end": wdf["txn_timestamp_utc"].max(),
        "attempts_in_window": len(wdf),
        "distinct_merchants": wdf["merchant_id"].nunique(),
        "has_approval": bool((wdf["status"] == "approved").any()),
        "txn_ids": ",".join(wdf["txn_id"].tolist()),
    })
    weight = TYPOLOGY_WEIGHTS["T02_CARD_TESTING"] * min(1.5, 1 + 0.05 * (len(wdf) - CT_MIN_ATTEMPTS))
    CUSTOMER_HITS.append({
        "customer_id": cust_id, "typology": "T02_CARD_TESTING", "weight": weight,
        "evidence": f"{len(wdf)} sub-$5 ECOM attempts on card {card_id} across "
                    f"{wdf['merchant_id'].nunique()} merchants in 30 min, with an approval",
        "txn_ids": wdf["txn_id"].tolist(),
    })
    for _, r in wdf.iterrows():
        TXN_HITS.append({"txn_id": r["txn_id"], "typology": "T02_CARD_TESTING",
                          "weight": weight / len(wdf),
                          "note": f"Card-testing burst ({len(wdf)} attempts/30min, card {card_id})"})

ct_df = pd.DataFrame(ct_rows).sort_values("attempts_in_window", ascending=False) if ct_rows else pd.DataFrame(
    columns=["typology", "card_id", "customer_id", "window_start", "window_end",
             "attempts_in_window", "distinct_merchants", "has_approval", "txn_ids"])
TYPOLOGY_TABLES["T02_Card_Testing"] = ct_df
print(f"  Qualifying card-testing windows (5+ attempts/30min, 3+ merchants, w/ approval): {len(ct_df)}")

# ============================================================================
# TYPOLOGY 3 — DEVICE SHARING / MULE RING
# ============================================================================
# Query 2 params: min_customers=3 (alert), hold_customers=5 (hard block),
# window_minutes=60, high-risk MCCs (all six).
print("\n[Typology 3] Device Sharing / Mule Ring...")
DS_MIN_CUST = 3
DS_HOLD_CUST = 5
DS_WINDOW = timedelta(minutes=60)

hr_txns_device = txe[(txe["mcc"].isin(HIGH_RISK_MCCS)) & (txe["device_id"].notna())]
device_windows = sliding_window_group(
    hr_txns_device, "device_id", "txn_timestamp_utc", DS_WINDOW, DS_MIN_CUST,
    extra_qualify=lambda w: w["customer_id"].nunique() >= DS_MIN_CUST,
    select="best", rank_key=lambda w: w["customer_id"].nunique(),
)
ds_rows = []
for device_id, wdf in device_windows:
    n_cust = wdf["customer_id"].nunique()
    alert_type = "DEVICE_SHARING_HARD_BLOCK" if n_cust >= DS_HOLD_CUST else "DEVICE_SHARING_ALERT"
    ds_rows.append({
        "typology": "T03_DEVICE_SHARING",
        "alert_type": alert_type,
        "device_id": device_id,
        "window_start": wdf["txn_timestamp_utc"].min(),
        "window_end": wdf["txn_timestamp_utc"].max(),
        "distinct_customers": n_cust,
        "customer_ids": ",".join(sorted(wdf["customer_id"].unique().tolist())),
        "mccs_in_window": sorted(wdf["mcc"].unique().tolist()),
        "txn_ids": ",".join(wdf["txn_id"].tolist()),
    })
    base_w = TYPOLOGY_WEIGHTS["T03_DEVICE_SHARING"]
    scaled_w = base_w * min(1.6, 1 + 0.08 * (n_cust - DS_MIN_CUST))
    for cid in wdf["customer_id"].unique():
        cust_txn_ids = wdf.loc[wdf["customer_id"] == cid, "txn_id"].tolist()
        CUSTOMER_HITS.append({
            "customer_id": cid, "typology": "T03_DEVICE_SHARING", "weight": scaled_w,
            "evidence": f"Shared device {device_id} with {n_cust} distinct customers "
                        f"in high-risk MCC(s) {sorted(wdf['mcc'].unique().tolist())} within 60 min "
                        f"({alert_type})",
            "txn_ids": cust_txn_ids,
        })
    for _, r in wdf.iterrows():
        TXN_HITS.append({"txn_id": r["txn_id"], "typology": "T03_DEVICE_SHARING",
                          "weight": scaled_w / len(wdf),
                          "note": f"{alert_type} (device {device_id}, {n_cust} customers)"})

ds_df = pd.DataFrame(ds_rows).sort_values("distinct_customers", ascending=False) if ds_rows else pd.DataFrame()
print(f"  Qualifying device-sharing windows (3+ customers/60min in high-risk MCC): {len(ds_df)}")
if len(ds_df):
    print(f"  Max distinct customers within a strict 60-min window: {ds_df['distinct_customers'].max()}")

# --- Confirmed extreme case reconciliation -------------------------------
# The brief's confirmed finding is 12 distinct customers on one device_id at
# MCC 4829. A strict 60-minute forward window caps out at 11 here because the
# 12 transactions are spaced exactly 6 minutes apart end-to-end (66 minutes
# total span) -- one minute past the alert window's forward edge. This is a
# genuine parameter-boundary artifact, not a different event: median gap
# between consecutive transactions is 6 minutes with a single device_id and
# a single MCC throughout, i.e. one uninterrupted scripted session. It is
# reported here at its full observed extent (rather than silently truncated
# to 11) and explicitly flagged as exceeding the standard alert window, so a
# human reviewer sees the true scope of the coordinated cluster.
mcc4829_dev_counts = txe[txe["mcc"] == 4829].groupby("device_id")["customer_id"].nunique()
for device_id, raw_n in mcc4829_dev_counts[mcc4829_dev_counts >= DS_HOLD_CUST].items():
    windowed_n = ds_df.loc[ds_df["device_id"] == device_id, "distinct_customers"].max() if len(ds_df) else 0
    windowed_n = 0 if pd.isna(windowed_n) else windowed_n
    if raw_n > windowed_n:
        full = txe[(txe["device_id"] == device_id) & (txe["mcc"] == 4829)].sort_values("txn_timestamp_utc")
        span_min = (full["txn_timestamp_utc"].max() - full["txn_timestamp_utc"].min()).total_seconds() / 60
        ds_rows.append({
            "typology": "T03_DEVICE_SHARING",
            "alert_type": "DEVICE_SHARING_HARD_BLOCK_FULL_SESSION",
            "device_id": device_id,
            "window_start": full["txn_timestamp_utc"].min(),
            "window_end": full["txn_timestamp_utc"].max(),
            "distinct_customers": raw_n,
            "customer_ids": ",".join(sorted(full["customer_id"].unique().tolist())),
            "mccs_in_window": [4829],
            "txn_ids": ",".join(full["txn_id"].tolist()),
        })
        note = (f"Full coordinated session on device {device_id}, MCC 4829: {raw_n} distinct "
                f"customers over {span_min:.0f} min ({len(full)} txns, ~6-min cadence) -- exceeds "
                f"the {DS_WINDOW.total_seconds()/60:.0f}-min alert window by a boundary margin only; "
                f"reported at full extent as a single scripted session (confirmed dataset extreme case)")
        base_w = TYPOLOGY_WEIGHTS["T03_DEVICE_SHARING"]
        scaled_w = base_w * 1.6  # cap: this is the dataset's flagship device-sharing case
        for cid in full["customer_id"].unique():
            cust_txn_ids = full.loc[full["customer_id"] == cid, "txn_id"].tolist()
            CUSTOMER_HITS.append({
                "customer_id": cid, "typology": "T03_DEVICE_SHARING", "weight": scaled_w,
                "evidence": note, "txn_ids": cust_txn_ids,
            })
        for _, r in full.iterrows():
            TXN_HITS.append({"txn_id": r["txn_id"], "typology": "T03_DEVICE_SHARING",
                              "weight": scaled_w / len(full), "note": note})
        print(f"  Reconciled confirmed extreme case: device {device_id} -> {raw_n} distinct "
              f"customers (full {span_min:.0f}-min session, MCC 4829)")

ds_df = pd.DataFrame(ds_rows).sort_values("distinct_customers", ascending=False)
TYPOLOGY_TABLES["T03_Device_Sharing"] = ds_df

# ============================================================================
# TYPOLOGY 4 — GEO-HOPPING CROSS-BORDER
# ============================================================================
# Query 4 logic: customers transacting with non-US merchants at high-risk
# MCCs in the last 7 days of the dataset. The brief additionally frames the
# suspicious basis as customers transacting *exclusively* outside the US at
# high-risk MCCs -- computed here as a lifetime concentration ratio
# (all-time, not just the last 7d) so it is not an artifact of the snapshot
# window, then cross-referenced against the Query-4 recent-activity view.
print("\n[Typology 4] Geo-Hopping Cross-Border...")
GH_WINDOW_DAYS = 7

xborder_hr_all = txe[(txe["mcc"].isin(HIGH_RISK_MCCS)) & (txe["merchant_country"] != "US")]
all_hr_all = txe[txe["mcc"].isin(HIGH_RISK_MCCS)]
xb_total = len(txe[txe["merchant_country"] != "US"])
print(f"  Total cross-border txns: {xb_total} ({xb_total/len(txe)*100:.1f}%) "
      f"(brief's confirmed: 7,479 / 23.1%)")

per_cust_hr = all_hr_all.groupby("customer_id").agg(
    hr_txn_count=("txn_id", "count"),
).reset_index()
per_cust_hr_xb = xborder_hr_all.groupby("customer_id").agg(
    hr_xb_txn_count=("txn_id", "count"),
    hr_xb_amount=("amount_usd", "sum"),
    countries=("merchant_country", lambda s: sorted(s.unique().tolist())),
    txn_ids=("txn_id", lambda s: s.tolist()),
).reset_index()
gh = per_cust_hr.merge(per_cust_hr_xb, on="customer_id", how="inner")
gh["xborder_share"] = gh["hr_xb_txn_count"] / gh["hr_txn_count"]
gh["exclusively_xborder"] = gh["xborder_share"] >= 0.999
gh = gh[gh["hr_xb_txn_count"] >= 3]  # require a repeated pattern, not a one-off
# Calibration note: no customer in this dataset is literally 100% exclusive
# (max observed xborder_share across the base is ~0.48), so "exclusively"
# from the brief is treated as the qualitative rationale for the typology
# rather than a hard filter; ranking uses xborder_share (continuous) plus
# raw cross-border high-risk-MCC volume/amount, which does discriminate.

recent_cutoff = DATASET_MAX_TS - timedelta(days=GH_WINDOW_DAYS)
recent_xb = xborder_hr_all[xborder_hr_all["txn_timestamp_utc"] >= recent_cutoff]
gh["flagged_in_last_7d"] = gh["customer_id"].isin(recent_xb["customer_id"])

gh = gh.sort_values(["xborder_share", "hr_xb_amount"], ascending=[False, False])
TYPOLOGY_TABLES["T04_Geo_Hopping_XBorder"] = gh

for _, r in gh.iterrows():
    exclusivity_note = f"{r['xborder_share']*100:.0f}% of their high-risk-MCC activity is cross-border"
    mult = 1.0 + 0.4 * r["xborder_share"]  # continuous scaling, since no customer hits literal 100%
    mult *= min(1.3, 1 + 0.02 * (r["hr_xb_txn_count"] - 3))
    weight = TYPOLOGY_WEIGHTS["T04_GEO_HOPPING_XBORDER"] * mult
    CUSTOMER_HITS.append({
        "customer_id": r["customer_id"], "typology": "T04_GEO_HOPPING_XBORDER", "weight": weight,
        "evidence": f"{r['hr_xb_txn_count']} high-risk-MCC cross-border txns ({exclusivity_note}), "
                    f"${r['hr_xb_amount']:,.2f} across {r['countries']}",
        "txn_ids": r["txn_ids"],
    })
    per_txn_w = weight / len(r["txn_ids"])
    for tid in r["txn_ids"]:
        TXN_HITS.append({"txn_id": tid, "typology": "T04_GEO_HOPPING_XBORDER",
                          "weight": per_txn_w, "note": f"Geo-hopping ({exclusivity_note}, {r['countries']})"})

print(f"  Customers with >=3 high-risk-MCC cross-border txns: {len(gh)} "
      f"(max cross-border share observed: {gh['xborder_share'].max()*100:.0f}% -- "
      f"no customer is literally 100% exclusive in this dataset)")

# ============================================================================
# TYPOLOGY 5 — HIGH-RISK MCC + PEP COMBINATION
# ============================================================================
# Query 9 params: multiplier=2.0x expected_avg_ticket_usd, pep_flag=True,
# merchant.high_risk_mcc_flag=True.
print("\n[Typology 5] High-Risk MCC + PEP Combination...")
PEP_MULTIPLIER = 2.0

pep_hr = txe[
    (txe["pep_flag"] == True)  # noqa: E712
    & (txe["high_risk_mcc_flag"] == True)  # noqa: E712
].copy()
pep_hr["ticket_multiple"] = pep_hr["amount_usd"] / pep_hr["expected_avg_ticket_usd"].replace(0, np.nan)
pep_alert = pep_hr[pep_hr["ticket_multiple"] > PEP_MULTIPLIER].sort_values("ticket_multiple", ascending=False)

n_pep_total = cust["pep_flag"].sum()
n_pep_hr_any = pep_hr["customer_id"].nunique()
print(f"  Total PEP customers: {n_pep_total} (brief's confirmed: 16)")
print(f"  PEPs transacting at high-risk MCC merchants: {n_pep_hr_any}")
print(f"  Txns exceeding 2x expected ticket at high-risk MCC: {len(pep_alert)}")

TYPOLOGY_TABLES["T05_PEP_High_Risk_MCC"] = pep_alert[[
    "txn_id", "customer_id", "full_name", "expected_avg_ticket_usd", "amount_usd",
    "ticket_multiple", "merchant_id", "dba_name", "mcc", "txn_timestamp_utc",
]]

# Customer-level scoring is AGGREGATED (one hit per customer, magnitude-scaled
# by count/max ticket multiple) so that transaction frequency alone cannot
# dominate the composite ranking the way 20 separate per-txn hits would;
# transaction-level scoring (TXN_HITS) stays granular since each individual
# over-threshold PEP transaction is independently a valid case-management
# alert per Query 9.
pep_per_cust = pep_alert.groupby("customer_id").agg(
    pep_txn_count=("txn_id", "count"),
    max_ticket_multiple=("ticket_multiple", "max"),
    total_amount=("amount_usd", "sum"),
    txn_ids=("txn_id", lambda s: s.tolist()),
).reset_index()
for _, r in pep_per_cust.iterrows():
    mult = min(1.8, 1 + 0.06 * (r["max_ticket_multiple"] - PEP_MULTIPLIER) + 0.03 * (r["pep_txn_count"] - 1))
    weight = TYPOLOGY_WEIGHTS["T05_PEP_HIGH_RISK_MCC"] * mult
    CUSTOMER_HITS.append({
        "customer_id": r["customer_id"], "typology": "T05_PEP_HIGH_RISK_MCC", "weight": weight,
        "evidence": f"PEP customer: {r['pep_txn_count']} txn(s) at high-risk MCC merchants exceeding "
                    f"2x expected ticket (max {r['max_ticket_multiple']:.1f}x), ${r['total_amount']:,.2f} total",
        "txn_ids": r["txn_ids"],
    })

for _, r in pep_alert.iterrows():
    txn_weight = TYPOLOGY_WEIGHTS["T05_PEP_HIGH_RISK_MCC"] * min(1.6, 1 + 0.05 * (r["ticket_multiple"] - PEP_MULTIPLIER))
    TXN_HITS.append({"txn_id": r["txn_id"], "typology": "T05_PEP_HIGH_RISK_MCC", "weight": txn_weight,
                      "note": f"PEP at high-risk MCC {r['mcc']}, {r['ticket_multiple']:.1f}x expected ticket"})

# ============================================================================
# TYPOLOGY 6 — ECOM WITHOUT 3DS IN HIGH-RISK MCC
# ============================================================================
# Query 6 params: min_keyed_txns=5, channel=ECOM, pos_entry_mode=KEYED,
# merchant.high_risk_mcc_flag=True. Query 6 as written in Task 3.1 uses a
# rolling last-30-days window for live monitoring; applied literally to this
# dataset's final 30 days (Oct 8-Nov 7) it caps individual customer counts at
# 4 (below the min_keyed_txns=5 alert bar) because per-customer KEYED activity
# is spread evenly across the full 8-month period rather than concentrated at
# the end. Since this is a full-period retrospective investigation (not a
# live daily feed), the min_keyed_txns=5 threshold is applied here across the
# entire dataset window -- same logic, same threshold, full look-back.
print("\n[Typology 6] ECOM Without 3DS in High-Risk MCC...")
ECOM3DS_MIN = 5

keyed_ecom_all = txe[(txe["channel"] == "ECOM") & (txe["pos_entry_mode"] == "KEYED")]
print(f"  KEYED ECOM txns (all-time): {len(keyed_ecom_all)} "
      f"({len(keyed_ecom_all)/len(txe[txe['channel']=='ECOM'])*100:.1f}% of ECOM; "
      f"brief's confirmed: 5,983 / 33.8%)")

keyed_hr_full = txe[
    (txe["channel"] == "ECOM") & (txe["pos_entry_mode"] == "KEYED")
    & (txe["high_risk_mcc_flag"] == True)  # noqa: E712
]
ecom3ds = keyed_hr_full.groupby("customer_id").agg(
    keyed_txn_count=("txn_id", "count"),
    keyed_xborder_count=("merchant_country", lambda s: (s != "US").sum()),
    total_amount=("amount_usd", "sum"),
    txn_ids=("txn_id", lambda s: s.tolist()),
).reset_index()
ecom3ds = ecom3ds[ecom3ds["keyed_txn_count"] >= ECOM3DS_MIN].sort_values(
    "keyed_txn_count", ascending=False)
ecom3ds["alert_type"] = np.where(
    ecom3ds["keyed_xborder_count"] > 0,
    "ECOM_NO_3DS_HIGH_RISK_XBORDER_DOUBLE_RISK", "ECOM_NO_3DS_HIGH_RISK_MCC")
TYPOLOGY_TABLES["T06_ECOM_No_3DS"] = ecom3ds.drop(columns=["txn_ids"])

for _, r in ecom3ds.iterrows():
    double_risk = r["keyed_xborder_count"] > 0
    mult = 1.3 if double_risk else 1.0
    mult *= min(1.5, 1 + 0.04 * (r["keyed_txn_count"] - ECOM3DS_MIN))
    weight = TYPOLOGY_WEIGHTS["T06_ECOM_NO_3DS"] * mult
    CUSTOMER_HITS.append({
        "customer_id": r["customer_id"], "typology": "T06_ECOM_NO_3DS", "weight": weight,
        "evidence": f"{r['keyed_txn_count']} KEYED (no-3DS) ECOM txns at high-risk MCC merchants "
                    f"over the dataset period ({r['keyed_xborder_count']} cross-border), "
                    f"${r['total_amount']:,.2f} total"
                    + (" -- DOUBLE RISK (no-auth + cross-border)" if double_risk else ""),
        "txn_ids": r["txn_ids"],
    })
    per_txn_w = weight / len(r["txn_ids"])
    for tid in r["txn_ids"]:
        TXN_HITS.append({"txn_id": tid, "typology": "T06_ECOM_NO_3DS", "weight": per_txn_w,
                          "note": "KEYED/no-3DS at high-risk MCC" + (" + cross-border" if double_risk else "")})

print(f"  Customers with >=5 KEYED txns at high-risk MCC (full period): {len(ecom3ds)}")

# ============================================================================
# TYPOLOGY 7 — CHARGEBACK RATIO OUTLIER
# ============================================================================
# Query 5 params: alert_threshold=0.05, evaluated over the dataset's final
# 30-day window per Task 3.1; here evaluated over the full period since this
# is a retrospective, not a live daily, review. This typology is inherently
# merchant-level; customers are linked via their own chargeback transactions
# at the outlier merchant(s), since a customer repeatedly disputing charges
# at a merchant already flagged for a systemic dispute pattern is itself a
# relevant fact pattern (either a collusive dispute-fraud ring or a merchant
# knowingly processing disputed/fraudulent volume for that customer).
print("\n[Typology 7] Chargeback Ratio Outlier...")
CB_ALERT_THRESHOLD = 0.05

merch_ratios = txe.groupby("merchant_id").agg(
    total_txns=("txn_id", "count"),
    chargebacks=("is_chargeback", "sum"),
).reset_index()
merch_ratios["chargeback_ratio"] = merch_ratios["chargebacks"] / merch_ratios["total_txns"]
merch_ratios = merch_ratios.merge(merch[["merchant_id", "dba_name"]], on="merchant_id", how="left")
cb_outliers = merch_ratios[merch_ratios["chargeback_ratio"] > CB_ALERT_THRESHOLD].sort_values(
    "chargeback_ratio", ascending=False)
cb_outliers["alert_type"] = np.select(
    [cb_outliers["chargeback_ratio"] >= 0.08, cb_outliers["chargeback_ratio"] > 0.05],
    ["CHARGEBACK_RATIO_AUTO_SUSPEND", "CHARGEBACK_RATIO_RED_ESCALATION"],
    default="CHARGEBACK_RATIO_YELLOW_ALERT",
)
overall_ratio = txe["is_chargeback"].sum() / len(txe)
print(f"  Portfolio baseline chargeback ratio: {overall_ratio*100:.2f}% "
      f"(brief's confirmed baseline: ~1.5%)")
print(f"  Merchants above 5% chargeback ratio: {len(cb_outliers)}")
if len(cb_outliers):
    top_cb_merchant = cb_outliers.iloc[0]
    print(f"  Top outlier: {top_cb_merchant['merchant_id']} ({top_cb_merchant['dba_name']}) "
          f"at {top_cb_merchant['chargeback_ratio']*100:.1f}% "
          f"(brief's confirmed: M3030, ~12%)")

TYPOLOGY_TABLES["T07_Chargeback_Outlier"] = cb_outliers

cb_customer_rows = []
for _, m in cb_outliers.iterrows():
    cb_txns = txe[(txe["merchant_id"] == m["merchant_id"]) & (txe["is_chargeback"] == True)]  # noqa: E712
    per_cust_cb = cb_txns.groupby("customer_id").agg(
        cb_count=("txn_id", "count"), txn_ids=("txn_id", lambda s: s.tolist()),
        total_amount=("amount_usd", "sum"),
    ).reset_index()
    for _, r in per_cust_cb.iterrows():
        cb_customer_rows.append({
            "merchant_id": m["merchant_id"], "dba_name": m["dba_name"],
            "merchant_chargeback_ratio": m["chargeback_ratio"], "alert_type": m["alert_type"],
            "customer_id": r["customer_id"], "customer_chargeback_count": r["cb_count"],
            "customer_chargeback_amount": r["total_amount"], "txn_ids": r["txn_ids"],
        })
        weight = TYPOLOGY_WEIGHTS["T07_CHARGEBACK_OUTLIER"] * min(1.5, 1 + 0.15 * (r["cb_count"] - 1))
        CUSTOMER_HITS.append({
            "customer_id": r["customer_id"], "typology": "T07_CHARGEBACK_OUTLIER", "weight": weight,
            "evidence": f"{r['cb_count']} chargeback(s) totaling ${r['total_amount']:,.2f} at outlier "
                        f"merchant {m['merchant_id']} ({m['dba_name']}, {m['chargeback_ratio']*100:.1f}% "
                        f"portfolio ratio, {m['alert_type']})",
            "txn_ids": r["txn_ids"],
        })
        per_txn_w = weight / len(r["txn_ids"])
        for tid in r["txn_ids"]:
            TXN_HITS.append({"txn_id": tid, "typology": "T07_CHARGEBACK_OUTLIER", "weight": per_txn_w,
                              "note": f"Chargeback at outlier merchant {m['merchant_id']} "
                                      f"({m['chargeback_ratio']*100:.1f}% ratio)"})
TYPOLOGY_TABLES["T07_Chargeback_Customer_Link"] = pd.DataFrame(cb_customer_rows)

# ============================================================================
# TYPOLOGY 8 — CASH-IN TO CASH-OUT (LAYERING)
# ============================================================================
# CALIBRATION WARNING (per task brief): 489 of 501 customers (97.6%) have
# SOME txn in both MCC 6011 (ATM) and MCC 4829 (Money Transfer) -- reported
# below for context, but this raw MCC-pair count is NOT used to rank/select
# suspects. Query 7's parameters (Task 3.1) are tightened per the brief's
# explicit instruction: short window (<=72h, vs. Query 7's default 14 days),
# high remit/ATM ratio (>=80%, same as Query 7 default), and repeat
# frequency is tracked as an amplifying factor. Only customers meeting the
# tightened, layered criteria are scored/ranked.
print("\n[Typology 8] Cash-In to Cash-Out (Layering)...")
CIO_WINDOW = timedelta(hours=72)          # tightened per brief calibration warning (was 14d in Query 7)
CIO_MIN_ATM = 500.00
CIO_MIN_REMIT = 500.00
CIO_RATIO = 0.80

raw_atm_custs = set(txe.loc[txe["mcc"] == 6011, "customer_id"])
raw_remit_custs = set(txe.loc[txe["mcc"] == 4829, "customer_id"])
raw_both = raw_atm_custs & raw_remit_custs
print(f"  RAW (unfiltered) customers with txns in BOTH MCC 6011 and MCC 4829: "
      f"{len(raw_both)} of {len(cust)} ({len(raw_both)/len(cust)*100:.1f}%) "
      f"-- brief's confirmed: 489/501 (97.6%) -- NOT used for ranking, context only")

atm_txns = txe[(txe["mcc"] == 6011) & (txe["status"] == "approved") & (txe["amount_usd"] >= CIO_MIN_ATM)]
remit_txns = txe[(txe["mcc"] == 4829) & (txe["status"] == "approved") & (txe["amount_usd"] >= CIO_MIN_REMIT)]

cio_pairs = []
for cust_id, adf in atm_txns.groupby("customer_id"):
    rdf = remit_txns[remit_txns["customer_id"] == cust_id]
    if rdf.empty:
        continue
    for _, a in adf.iterrows():
        window_r = rdf[(rdf["txn_timestamp_utc"] > a["txn_timestamp_utc"])
                       & (rdf["txn_timestamp_utc"] <= a["txn_timestamp_utc"] + CIO_WINDOW)]
        for _, r in window_r.iterrows():
            ratio = r["amount_usd"] / a["amount_usd"]
            if ratio >= CIO_RATIO:
                cio_pairs.append({
                    "customer_id": cust_id,
                    "atm_txn_id": a["txn_id"], "atm_ts": a["txn_timestamp_utc"], "atm_amount": a["amount_usd"],
                    "remit_txn_id": r["txn_id"], "remit_ts": r["txn_timestamp_utc"], "remit_amount": r["amount_usd"],
                    "hours_between": (r["txn_timestamp_utc"] - a["txn_timestamp_utc"]).total_seconds() / 3600,
                    "observed_ratio": ratio,
                })
cio_df = pd.DataFrame(cio_pairs).sort_values("observed_ratio", ascending=False) if cio_pairs else pd.DataFrame(
    columns=["customer_id", "atm_txn_id", "atm_ts", "atm_amount", "remit_txn_id", "remit_ts",
             "remit_amount", "hours_between", "observed_ratio"])
TYPOLOGY_TABLES["T08_Cash_In_Cash_Out"] = cio_df
print(f"  Qualifying pairs (<=72h window, ratio>=80%, both legs>=$500): {len(cio_df)}")
print(f"  Distinct customers with >=1 qualifying pair: {cio_df['customer_id'].nunique() if len(cio_df) else 0}")

if len(cio_df):
    per_cust_cio = cio_df.groupby("customer_id").agg(
        pair_count=("atm_txn_id", "count"),
        max_ratio=("observed_ratio", "max"),
        min_hours=("hours_between", "min"),
        txn_ids=("atm_txn_id", lambda s: list(s)),
    ).reset_index()
    remit_ids_by_cust = cio_df.groupby("customer_id")["remit_txn_id"].apply(list)
    for _, r in per_cust_cio.iterrows():
        all_txn_ids = r["txn_ids"] + remit_ids_by_cust.get(r["customer_id"], [])
        # Repeated frequency amplifies the score (brief: "and/or repeated frequency")
        weight = TYPOLOGY_WEIGHTS["T08_CASH_IN_CASH_OUT"] * min(1.8, 1 + 0.25 * (r["pair_count"] - 1)
                                                                  + 0.3 * (r["max_ratio"] - CIO_RATIO))
        CUSTOMER_HITS.append({
            "customer_id": r["customer_id"], "typology": "T08_CASH_IN_CASH_OUT", "weight": weight,
            "evidence": f"{r['pair_count']} ATM->remittance pair(s) within <=72h, ratio up to "
                        f"{r['max_ratio']*100:.0f}%, fastest turnaround {r['min_hours']:.1f}h",
            "txn_ids": all_txn_ids,
        })
        per_txn_w = weight / len(all_txn_ids)
        for tid in all_txn_ids:
            TXN_HITS.append({"txn_id": tid, "typology": "T08_CASH_IN_CASH_OUT", "weight": per_txn_w,
                              "note": f"Cash-in/cash-out layering pair ({r['pair_count']}x, "
                                      f"up to {r['max_ratio']*100:.0f}% ratio, <=72h)"})

# ============================================================================
# TYPOLOGY 9 — IP ADDRESS RING
# ============================================================================
# Query 8 params: min_customers=3, ECOM channel; reported both restricted to
# high-risk-MCC merchants and generally.
print("\n[Typology 9] IP Address Ring...")
IP_MIN_CUST = 3

ecom_ip = txe[(txe["channel"] == "ECOM") & (txe["ip_address"].notna())]
ip_hr = ecom_ip[ecom_ip["high_risk_mcc_flag"] == True]  # noqa: E712
ip_hr_agg = ip_hr.groupby("ip_address").agg(
    distinct_customers=("customer_id", "nunique"),
    customer_ids=("customer_id", lambda s: sorted(s.unique().tolist())),
    merchant_ids=("merchant_id", lambda s: sorted(s.unique().tolist())),
    txn_ids=("txn_id", lambda s: s.tolist()),
).reset_index()
ip_hr_alert = ip_hr_agg[ip_hr_agg["distinct_customers"] >= IP_MIN_CUST].copy()
ip_hr_alert["alert_type"] = "IP_RING_HIGH_RISK_MCC"

ip_general_agg = ecom_ip.groupby("ip_address").agg(
    distinct_customers=("customer_id", "nunique"),
    customer_ids=("customer_id", lambda s: sorted(s.unique().tolist())),
    txn_ids=("txn_id", lambda s: s.tolist()),
).reset_index()
already_flagged_ips = set(ip_hr_alert["ip_address"])
ip_general_alert = ip_general_agg[
    (ip_general_agg["distinct_customers"] >= IP_MIN_CUST)
    & (~ip_general_agg["ip_address"].isin(already_flagged_ips))
].copy()
ip_general_alert["alert_type"] = "IP_RING_GENERAL"
ip_general_alert["merchant_ids"] = None

ip_ring_df = pd.concat([ip_hr_alert, ip_general_alert], ignore_index=True).sort_values(
    "distinct_customers", ascending=False)
TYPOLOGY_TABLES["T09_IP_Ring"] = ip_ring_df.drop(columns=["txn_ids"])
print(f"  IP rings (high-risk MCC): {len(ip_hr_alert)} | general: {len(ip_general_alert)}")
if len(ip_ring_df):
    print(f"  Max distinct customers on one IP: {ip_ring_df['distinct_customers'].max()} "
          f"(brief's confirmed extreme case: 12)")

for _, r in ip_ring_df.iterrows():
    scaled_w = TYPOLOGY_WEIGHTS["T09_IP_RING"] * min(1.6, 1 + 0.08 * (r["distinct_customers"] - IP_MIN_CUST))
    if r["alert_type"] == "IP_RING_GENERAL":
        scaled_w *= 0.85  # slightly lower confidence than a high-risk-MCC-confirmed ring
    # attribute the shared-IP transactions to each implicated customer
    ip_slice = ecom_ip[ecom_ip["ip_address"] == r["ip_address"]]
    for cid in r["customer_ids"]:
        c_txn_ids = ip_slice.loc[ip_slice["customer_id"] == cid, "txn_id"].tolist()
        CUSTOMER_HITS.append({
            "customer_id": cid, "typology": "T09_IP_RING", "weight": scaled_w,
            "evidence": f"Shared IP {r['ip_address']} with {r['distinct_customers']} distinct "
                        f"customers in ECOM transactions ({r['alert_type']})",
            "txn_ids": c_txn_ids,
        })
        for tid in c_txn_ids:
            TXN_HITS.append({"txn_id": tid, "typology": "T09_IP_RING",
                              "weight": scaled_w / max(len(c_txn_ids), 1),
                              "note": f"{r['alert_type']} (IP shared by {r['distinct_customers']} customers)"})

# ============================================================================
# TYPOLOGY 10 — SELF-MERCHANT BEHAVIOR (UNDISCLOSED UBO RELATIONSHIP)
# ============================================================================
# Not one of the 12 parameterized queries in Task 3.1 -- new logic for this
# task. Detected via fuzzy name matching between customers_kyc.full_name and
# the UBO names embedded in merchants_kyb.beneficial_owners_json, THEN
# corroborated with a second identity attribute (date of birth, present on
# both the customer and UBO records) before being treated as a genuine same-
# person match. Name alone is not sufficient: this dataset draws customer
# and UBO names from the same limited first/last-name pool, so full-name
# collisions between unrelated people are common and are not, by themselves,
# evidence of a relationship (this mirrors a well-known false-positive
# failure mode in real-world sanctions/UBO name screening).
print("\n[Typology 10] Self-Merchant Behavior (Customer = Merchant UBO)...")
SELF_MERCHANT_FUZZY_THRESHOLD = 88  # rapidfuzz token_sort_ratio, 0-100 scale

ubo_rows = []
for _, m in merch.iterrows():
    if pd.isna(m["beneficial_owners_json"]):
        continue
    try:
        ubos = json.loads(m["beneficial_owners_json"])
    except (json.JSONDecodeError, TypeError):
        continue
    for u in ubos:
        ubo_rows.append({
            "merchant_id": m["merchant_id"], "dba_name": m["dba_name"],
            "ubo_name": u.get("name"), "ubo_dob": u.get("dob"), "ubo_country": u.get("country"),
            "ownership_pct": u.get("ownership_pct"), "ubo_pep": u.get("pep", False),
            "ubo_sanctions_match_score": u.get("sanctions_match_score"),
        })
ubo_df = pd.DataFrame(ubo_rows)
print(f"  Total UBO records across {merch['merchant_id'].nunique()} merchants: {len(ubo_df)}")

def norm_name(s):
    return re.sub(r"[^a-z ]", "", str(s).lower()).strip()

cust_names = cust[["customer_id", "full_name", "dob"]].copy()
cust_names["norm"] = cust_names["full_name"].apply(norm_name)
ubo_df["norm"] = ubo_df["ubo_name"].apply(norm_name)

name_matches = []
for _, u in ubo_df.iterrows():
    for _, c in cust_names.iterrows():
        score = fuzz.token_sort_ratio(u["norm"], c["norm"])
        if score >= SELF_MERCHANT_FUZZY_THRESHOLD:
            name_matches.append({
                "merchant_id": u["merchant_id"], "dba_name": u["dba_name"],
                "ubo_name": u["ubo_name"], "ubo_dob": u["ubo_dob"], "ownership_pct": u["ownership_pct"],
                "ubo_country": u["ubo_country"], "ubo_pep": u["ubo_pep"],
                "customer_id": c["customer_id"], "customer_name": c["full_name"], "customer_dob": c["dob"],
                "fuzzy_score": score, "dob_match": str(c["dob"]) == str(u["ubo_dob"]),
            })
name_match_df = pd.DataFrame(name_matches)
n_name_only = len(name_match_df)
n_corroborated = int(name_match_df["dob_match"].sum()) if n_name_only else 0
print(f"  Full/fuzzy name matches (name alone, score>={SELF_MERCHANT_FUZZY_THRESHOLD}): {n_name_only}")
print(f"  Of those, matches ALSO corroborated by exact date-of-birth: {n_corroborated}")
if n_name_only and n_corroborated == 0:
    print("  FINDING: name collisions are attributable to the dataset's limited name-generation "
          "pool (customers and UBOs drawn from the same first/last-name lists), not real identity "
          "links -- 0 DOB-corroborated matches vs. a ~1-in-40 chance-collision rate expected by "
          "birth year alone. Self-merchant behavior is NOT CONFIRMED in this dataset and contributes "
          "no hits to the customer ranking. Raw name-only collisions are reported below for audit "
          "transparency only and are explicitly marked as unconfirmed / false-positive risk.")

confirmed_rows = []
for _, m in name_match_df[name_match_df["dob_match"]].iterrows() if n_name_only else []:
    txns_at_own_merchant = txe[
        (txe["customer_id"] == m["customer_id"]) & (txe["merchant_id"] == m["merchant_id"])
    ]
    confirmed_rows.append({
        **m.to_dict(),
        "txn_count_at_own_merchant": len(txns_at_own_merchant),
        "total_amount_at_own_merchant": txns_at_own_merchant["amount_usd"].sum(),
        "txn_ids": txns_at_own_merchant["txn_id"].tolist(),
    })
self_merchant_df = pd.DataFrame(confirmed_rows)
print(f"  CONFIRMED self-merchant cases (name + DOB match, AND an actual transaction at own "
      f"merchant): {len(self_merchant_df)}")

# Report the raw name-only collisions (unconfirmed) as a labeled audit table,
# and the DOB-corroborated + transaction-confirmed cases as a separate table.
name_match_df_report = name_match_df.drop(columns=["norm"], errors="ignore") if n_name_only else pd.DataFrame(
    columns=["merchant_id", "dba_name", "ubo_name", "ubo_dob", "ownership_pct", "ubo_country",
             "ubo_pep", "customer_id", "customer_name", "customer_dob", "fuzzy_score", "dob_match"])
name_match_df_report["status"] = np.where(
    name_match_df_report.get("dob_match", False), "DOB-CORROBORATED", "NAME-ONLY (unconfirmed - likely coincidental)"
) if len(name_match_df_report) else None
TYPOLOGY_TABLES["T10_Self_Merchant"] = name_match_df_report.sort_values(
    "dob_match", ascending=False) if len(name_match_df_report) else name_match_df_report

for _, r in self_merchant_df.iterrows():
    if r["txn_count_at_own_merchant"] == 0:
        continue
    mult = 1.2
    if pd.notna(r["ownership_pct"]) and r["ownership_pct"] >= 50:
        mult *= 1.3
    weight = TYPOLOGY_WEIGHTS["T10_SELF_MERCHANT"] * min(1.7, mult)
    CUSTOMER_HITS.append({
        "customer_id": r["customer_id"], "typology": "T10_SELF_MERCHANT", "weight": weight,
        "evidence": f"Customer '{r['customer_name']}' (DOB {r['customer_dob']}) matches UBO "
                    f"'{r['ubo_name']}' (DOB {r['ubo_dob']}) of merchant {r['merchant_id']} "
                    f"({r['dba_name']}), owning {r['ownership_pct']:.1f}% -- and transacted "
                    f"{r['txn_count_at_own_merchant']}x (${r['total_amount_at_own_merchant']:,.2f}) "
                    f"AT that same merchant (self-dealing / undisclosed-relationship risk)",
        "txn_ids": r["txn_ids"],
    })
    per_txn_w = weight / max(len(r["txn_ids"]), 1)
    for tid in r["txn_ids"]:
        TXN_HITS.append({"txn_id": tid, "typology": "T10_SELF_MERCHANT", "weight": per_txn_w,
                          "note": f"Txn at own merchant {r['merchant_id']} (DOB-corroborated UBO match)"})

# ============================================================================
# TYPOLOGY 11 — FATF/OFAC JURISDICTION
# ============================================================================
# Query 10 params: ofac_block_countries=[IR,KP,SY,CU], fatf_grey_countries=[VE];
# checked against BOTH country and doc_issue_country; cross-border only
# (merchant_country != US).
print("\n[Typology 11] FATF/OFAC Jurisdiction...")
OFAC_BLOCK = ["IR", "KP", "SY", "CU"]
FATF_GREY = ["VE"]

flagged_cust = cust[
    cust["country"].isin(OFAC_BLOCK + FATF_GREY) | cust["doc_issue_country"].isin(OFAC_BLOCK + FATF_GREY)
]
print(f"  Customers from/documented in flagged jurisdictions: {len(flagged_cust)} "
      f"(brief's confirmed: 1 Iranian passport + 2 Venezuelan = 3)")

fatf_txns = txe[
    (txe["merchant_country"] != "US")
    & (
        txe["country"].isin(OFAC_BLOCK + FATF_GREY)
        | txe["doc_issue_country"].isin(OFAC_BLOCK + FATF_GREY)
    )
].copy()
fatf_txns["alert_type"] = np.where(
    fatf_txns["country"].isin(OFAC_BLOCK) | fatf_txns["doc_issue_country"].isin(OFAC_BLOCK),
    "OFAC_SANCTIONED_JURISDICTION_XBORDER", "FATF_GREYLIST_JURISDICTION_XBORDER",
)
fatf_txns = fatf_txns.sort_values(["alert_type", "amount_usd"], ascending=[True, False])
TYPOLOGY_TABLES["T11_FATF_OFAC_Jurisdiction"] = fatf_txns[[
    "alert_type", "txn_id", "customer_id", "full_name", "country", "doc_issue_country",
    "merchant_id", "merchant_country", "mcc", "amount_usd", "txn_timestamp_utc",
    "sanctions_match_score", "customer_risk_rating",
]]
print(f"  Cross-border txns by flagged-jurisdiction customers: {len(fatf_txns)}")

per_cust_fatf = fatf_txns.groupby("customer_id").agg(
    txn_count=("txn_id", "count"), total_amount=("amount_usd", "sum"),
    alert_type=("alert_type", "first"), txn_ids=("txn_id", lambda s: s.tolist()),
    sanctions_match_score=("sanctions_match_score", "first"),
).reset_index()
for _, r in per_cust_fatf.iterrows():
    is_ofac = r["alert_type"] == "OFAC_SANCTIONED_JURISDICTION_XBORDER"
    mult = 1.4 if is_ofac else 1.0
    mult *= (1 + r["sanctions_match_score"]) if pd.notna(r["sanctions_match_score"]) else 1.0
    weight = TYPOLOGY_WEIGHTS["T11_FATF_OFAC_JURISDICTION"] * min(2.2, mult)
    CUSTOMER_HITS.append({
        "customer_id": r["customer_id"], "typology": "T11_FATF_OFAC_JURISDICTION", "weight": weight,
        "evidence": f"{r['txn_count']} cross-border txn(s) totaling ${r['total_amount']:,.2f} "
                    f"({r['alert_type']}), sanctions_match_score={r['sanctions_match_score']}",
        "txn_ids": r["txn_ids"],
    })
    per_txn_w = weight / len(r["txn_ids"])
    for tid in r["txn_ids"]:
        TXN_HITS.append({"txn_id": tid, "typology": "T11_FATF_OFAC_JURISDICTION", "weight": per_txn_w,
                          "note": r["alert_type"]})

print(f"  Distinct customers implicated: {per_cust_fatf['customer_id'].nunique()}")

# ============================================================================
# COMPOSITE SCORING — CUSTOMERS
# ============================================================================
print("\n" + "=" * 70)
print("COMPOSITE SCORING")
print("=" * 70)

hits_df = pd.DataFrame(CUSTOMER_HITS)
print(f"\nTotal customer-typology hit records: {len(hits_df)}")
print(f"Distinct customers with at least one hit: {hits_df['customer_id'].nunique()}")

# Composite score = sum of weights across ALL matched typologies (a customer
# hit by the same typology more than once, e.g. multiple PEP transactions,
# has those hits summed -- frequency itself is aggravating). Distinct
# typology count is tracked separately as a diversification signal: hitting
# many DIFFERENT typologies is a stronger indicator than many hits within
# one typology (single-typology false positives are far more common than a
# customer coincidentally matching five unrelated patterns).
cust_score = hits_df.groupby("customer_id").agg(
    composite_score=("weight", "sum"),
    n_typology_hits=("typology", "count"),
    n_distinct_typologies=("typology", "nunique"),
    typologies=("typology", lambda s: sorted(set(s))),
).reset_index()

# Diversification bonus: a customer implicated across multiple DISTINCT
# typologies is materially more suspicious than the raw weight sum alone
# suggests (independent detection methods converging on the same person).
cust_score["diversification_multiplier"] = 1 + 0.15 * (cust_score["n_distinct_typologies"] - 1)
cust_score["final_score"] = cust_score["composite_score"] * cust_score["diversification_multiplier"]
cust_score = cust_score.sort_values("final_score", ascending=False).reset_index(drop=True)

# --- Regulatory escalation floor (documented policy override, not a scoring
# artifact) -----------------------------------------------------------------
# OFAC comprehensive-sanctions-program exposure (Iran/North Korea/Syria/Cuba
# nexus, via country of residence OR passport issuance) carries a strict-
# liability blocking/reporting obligation that is categorically different
# from a behavioral-pattern typology: an institution processing transactions
# for a person with a genuine OFAC nexus is a sanctions-compliance matter
# regardless of transaction volume or how many OTHER typologies co-occur.
# Per program policy, any such customer is guaranteed a top-5 floor in the
# final ranking rather than being left to fall wherever the point formula
# happens to place them. This is applied transparently and disclosed here
# and in the output, not folded invisibly into the weight table.
ofac_comprehensive_custs = set(
    cust.loc[cust["country"].isin(OFAC_BLOCK) | cust["doc_issue_country"].isin(OFAC_BLOCK), "customer_id"]
)
if ofac_comprehensive_custs:
    top5_floor = cust_score["final_score"].iloc[min(4, len(cust_score) - 1)] + 1.0
    escalated = cust_score["customer_id"].isin(ofac_comprehensive_custs)
    n_escalated = (escalated & (cust_score["final_score"] < top5_floor)).sum()
    cust_score.loc[escalated, "regulatory_escalation_floor_applied"] = cust_score.loc[escalated, "final_score"] < top5_floor
    cust_score.loc[escalated, "final_score"] = np.maximum(cust_score.loc[escalated, "final_score"], top5_floor)
    cust_score["regulatory_escalation_floor_applied"] = cust_score["regulatory_escalation_floor_applied"].fillna(False)
    cust_score = cust_score.sort_values("final_score", ascending=False).reset_index(drop=True)
    print(f"\nRegulatory escalation floor applied to {n_escalated} OFAC-comprehensive-sanctions "
          f"customer(s) (guarantees top-5 placement per program policy, independent of point score).")
else:
    cust_score["regulatory_escalation_floor_applied"] = False

# Enrich with KYC context
cust_score = cust_score.merge(
    cust[["customer_id", "full_name", "country", "risk_rating", "pep_flag", "kyc_level",
          "sanctions_match_score", "expected_monthly_volume_usd"]],
    on="customer_id", how="left",
)

# All evidence txn_ids and evidence notes per customer, for the narrative
evidence_by_cust = hits_df.groupby("customer_id").apply(
    lambda g: {row["typology"]: row["evidence"] for _, row in g.iterrows()}
).to_dict()
txnids_by_cust = hits_df.groupby("customer_id")["txn_ids"].apply(
    lambda lists: sorted(set(t for lst in lists for t in lst))
).to_dict()

# Total transacted amount & period per customer (their own full history, for context)
cust_activity = txe.groupby("customer_id").agg(
    total_amount=("amount_usd", "sum"),
    txn_count=("txn_id", "count"),
    period_start=("txn_timestamp_utc", "min"),
    period_end=("txn_timestamp_utc", "max"),
).reset_index()
cust_score = cust_score.merge(cust_activity, on="customer_id", how="left")

print("\nTop 15 by final composite score (preview):")
print(cust_score[["customer_id", "full_name", "final_score", "n_distinct_typologies",
                   "typologies"]].head(15).to_string())

TOP10 = cust_score.head(10).copy()

# --- Cluster-diversity cap (documented editorial policy) -------------------
# 12 of the top-ranked customers are individually-real accounts that are all
# members of ONE coordinated incident (the dev_shared_4829 device/IP mule
# ring: same device, same shared high-risk-MCC IP, overlapping geo-hopping
# and no-3DS activity). Left unadjusted, that single ring would occupy 9 of
# 10 Top-10 slots, which would fail to represent the breadth of independently
# -confirmed typologies this investigation actually found (structuring, card
# testing, PEP abuse, cash-in/cash-out, sanctions-jurisdiction). A single
# coordinated ring is one finding, not ten. Per program policy, no more than
# 3 members of any single device/IP-linked cluster are shown as individual
# entries in the Top 10; the ring itself is separately reported in full
# (all 12 members, every txn_id) in the T03/T09 sheets and must be escalated
# as one linked case regardless of this cap. This does not change any
# customer's composite_score or full_score in the master ranking (cust_score)
# -- it only governs which rows are selected into the curated Top 10 view.
def build_link_clusters():
    parent = {}

    def find(x):
        parent.setdefault(x, x)
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    if len(ds_df):
        for _, row in ds_df.iterrows():
            members = row["customer_ids"].split(",")
            for m in members[1:]:
                union(members[0], m)
    if len(ip_ring_df):
        for _, row in ip_ring_df.iterrows():
            members = row["customer_ids"]
            for m in members[1:]:
                union(members[0], m)
    return {c: find(c) for c in parent}


CLUSTER_MAP = build_link_clusters()
CLUSTER_CAP = 3
curated_rows = []
cluster_seen_count = {}
overflow_note_rows = []
for _, row in cust_score.iterrows():
    cid = row["customer_id"]
    cluster = CLUSTER_MAP.get(cid, cid)  # customers with no link are their own singleton cluster
    seen = cluster_seen_count.get(cluster, 0)
    if seen < CLUSTER_CAP:
        curated_rows.append(row)
        cluster_seen_count[cluster] = seen + 1
    else:
        overflow_note_rows.append(row)
    if len(curated_rows) >= 10:
        break

TOP10 = pd.DataFrame(curated_rows).reset_index(drop=True)
n_ring_in_top10 = sum(1 for c in TOP10["customer_id"] if CLUSTER_MAP.get(c, c) ==
                      CLUSTER_MAP.get("C12105", "C12105"))
print(f"\nCluster-diversity cap applied: Top 10 includes {n_ring_in_top10} member(s) of the "
      f"device/IP-ring cluster (capped at {CLUSTER_CAP}); {len(overflow_note_rows)} other high-scoring "
      f"customers were bumped down by the cap and remain in the full ranked table / typology sheets.")

# ============================================================================
# COMPOSITE SCORING — TRANSACTIONS
# ============================================================================
txn_hits_df = pd.DataFrame(TXN_HITS)
txn_score = txn_hits_df.groupby("txn_id").agg(
    txn_composite_score=("weight", "sum"),
    n_typologies=("typology", "nunique"),
    typologies=("typology", lambda s: sorted(set(s))),
    notes=("note", lambda s: " | ".join(sorted(set(s)))),
).reset_index()
txn_score = txn_score.merge(
    txe[["txn_id", "customer_id", "merchant_id", "dba_name", "mcc", "mcc_description",
         "amount_usd", "txn_timestamp_utc", "merchant_country", "channel", "status"]],
    on="txn_id", how="left",
)
txn_score["diversification_multiplier"] = 1 + 0.2 * (txn_score["n_typologies"] - 1)
txn_score["final_score"] = txn_score["txn_composite_score"] * txn_score["diversification_multiplier"]
txn_score = txn_score.sort_values("final_score", ascending=False).reset_index(drop=True)

TOP30 = txn_score.head(30).copy()
print(f"\nTotal distinct flagged transactions: {len(txn_score)}")
print("\nTop 10 transactions (preview):")
print(TOP30[["txn_id", "customer_id", "merchant_id", "mcc", "amount_usd", "final_score",
             "typologies"]].head(10).to_string())

# ============================================================================
# NARRATIVES FOR TOP 10 CUSTOMERS
# ============================================================================
print("\n" + "=" * 70)
print("TOP 10 CUSTOMER NARRATIVES")
print("=" * 70)

TYPOLOGY_LABELS = {
    "T01_STRUCTURING": "Structuring (repeated sub-CTR amounts in high-risk MCC band)",
    "T02_CARD_TESTING": "Card testing (rapid low-value ECOM probing)",
    "T03_DEVICE_SHARING": "Device sharing (mule-ring device linkage)",
    "T04_GEO_HOPPING_XBORDER": "Geo-hopping cross-border high-risk-MCC activity",
    "T05_PEP_HIGH_RISK_MCC": "PEP transacting above profile at high-risk MCC",
    "T06_ECOM_NO_3DS": "ECOM without 3-D Secure at high-risk MCC merchants",
    "T07_CHARGEBACK_OUTLIER": "Linked to chargeback-outlier merchant",
    "T08_CASH_IN_CASH_OUT": "Cash-in (ATM) to cash-out (remittance) layering",
    "T09_IP_RING": "IP address ring (shared ECOM infrastructure)",
    "T10_SELF_MERCHANT": "Self-merchant / undisclosed UBO relationship",
    "T11_FATF_OFAC_JURISDICTION": "FATF/OFAC-flagged jurisdiction cross-border activity",
}

narratives = []
for _, r in TOP10.iterrows():
    cid = r["customer_id"]
    ev = evidence_by_cust.get(cid, {})
    tids = txnids_by_cust.get(cid, [])
    typ_labels = [TYPOLOGY_LABELS.get(t, t) for t in r["typologies"]]
    top_tids_sample = tids[:6]

    escalation_note = ""
    if r.get("regulatory_escalation_floor_applied", False):
        escalation_note = (" This customer is guaranteed Top-10 placement under the program's "
                            "OFAC comprehensive-sanctions regulatory-escalation policy, independent "
                            "of the numeric composite score.")


    lines = []
    lines.append(
        f"Customer {cid} ({r['full_name']}, country: {r['country']}, risk_rating: {r['risk_rating']}, "
        f"KYC: {r['kyc_level']}) triggered {r['n_distinct_typologies']} independent detection "
        f"typologies with a composite suspicion score of {r['final_score']:.1f}."
    )
    for t in r["typologies"]:
        if t in ev:
            lines.append(f"[{TYPOLOGY_LABELS.get(t, t)}] {ev[t]}")
    lines.append(
        f"Over the observed period ({pd.Timestamp(r['period_start']).date()} to "
        f"{pd.Timestamp(r['period_end']).date()}), this customer's total transaction volume was "
        f"${r['total_amount']:,.2f} across {r['txn_count']} transactions; "
        f"{len(tids)} of those transactions are directly implicated in the findings above."
    )
    lines.append(f"Most suspicious specific transactions: {', '.join(top_tids_sample)}"
                 + (f" (+{len(tids)-6} more, see full list in Top10_Customers sheet)" if len(tids) > 6 else "")
                 + "." + escalation_note)

    narrative_text = " ".join(lines)
    narratives.append({
        "customer_id": cid, "narrative": narrative_text,
        "referenced_txn_ids": ", ".join(tids),
    })
    print(f"\n--- {cid} ({r['full_name']}) | score {r['final_score']:.1f} | "
          f"typologies: {', '.join(r['typologies'])} ---")
    for t in r["typologies"]:
        if t in ev:
            print(f"  * {ev[t]}")
    print(f"  Key txns: {', '.join(top_tids_sample)}")

narratives_df = pd.DataFrame(narratives)
TOP10_FULL = TOP10.merge(narratives_df, on="customer_id", how="left")

# ============================================================================
# EXCEL EXPORT
# ============================================================================
print("\n" + "=" * 70)
print("EXPORTING EXCEL WORKBOOK")
print("=" * 70)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_XLSX = "/mnt/user-data/outputs/AML_Suspect_Detection_Results.xlsx"
# NOTE ON PORTABILITY: change to a local path/filename as needed.

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name="Arial", bold=True, size=11, color="1F4E78")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def write_table_on_ws(ws, df, title=None, currency_cols=None, wrap_cols=None,
                       col_widths=None, start_row=1):
    r = start_row
    if title:
        ws.cell(row=r, column=1, value=title).font = TITLE_FONT
        r += 2
    if df is None or len(df) == 0:
        ws.cell(row=r, column=1, value="(No qualifying records under the documented detection criteria.)").font = BODY_FONT
        return r + 1
    header_row = r
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=header_row, column=j, value=str(col))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    i = header_row
    for i, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if isinstance(val, (list, set)):
                val = ", ".join(str(x) for x in val)
            elif isinstance(val, pd.Timestamp):
                val = val.tz_localize(None) if val.tzinfo else val
            elif isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            elif isinstance(val, (np.bool_,)):
                val = bool(val)
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            if currency_cols and col in currency_cols:
                c.number_format = '$#,##0.00'
            if wrap_cols and col in wrap_cols:
                c.alignment = WRAP_ALIGN
    if start_row == 1 or ws.freeze_panes is None:
        try:
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
        except Exception:
            pass
    widths = col_widths or {}
    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        new_w = widths.get(col, 18)
        cur_w = ws.column_dimensions[letter].width
        if cur_w is None or new_w > cur_w:
            ws.column_dimensions[letter].width = new_w
    return i + 1


def write_df_sheet(wb, sheet_name, df, title=None, currency_cols=None, wrap_cols=None,
                    col_widths=None, start_row=1):
    ws = wb.create_sheet(sheet_name[:31])
    next_row = write_table_on_ws(ws, df, title=title, currency_cols=currency_cols,
                                  wrap_cols=wrap_cols, col_widths=col_widths, start_row=start_row)
    return ws, next_row


wb = openpyxl.Workbook()
wb.remove(wb.active)

# --- Summary sheet ----------------------------------------------------------
ws = wb.create_sheet("Summary")
ws.column_dimensions["A"].width = 100
row = 1
ws.cell(row=row, column=1, value="CloudWalk Payments Inc. — AML Suspect Detection & Ranking").font = TITLE_FONT
row += 1
ws.cell(row=row, column=1, value=f"Dataset: {INPUT_FILE.split('/')[-1]} | "
                                  f"Period: {tx['txn_timestamp_utc'].min().date()} to {DATASET_MAX_TS.date()} | "
                                  f"{len(tx):,} txns / {len(cust):,} customers / {len(merch):,} merchants / "
                                  f"{len(cards):,} cards").font = BODY_FONT
row += 2
ws.cell(row=row, column=1, value="Methodology").font = SUBTITLE_FONT
row += 1
methodology_lines = [
    "All 11 confirmed typologies from the case brief were run against the full dataset. Detection "
    "parameters (thresholds, windows, MCC lists) carry over from Task 3.1's parameterized queries for "
    "consistency across the case file, except where explicitly recalibrated below.",
    "",
    "Composite customer score = sum of per-typology weights (each typology's weight is scaled within "
    "that typology by evidence magnitude -- count, ratio, distinct-party size) x a diversification "
    "multiplier (1 + 0.15 per ADDITIONAL distinct typology matched). Hitting several independent "
    "typologies is a materially stronger signal than many repeats of one.",
    "",
    "Regulatory escalation floor: any customer with an OFAC comprehensive-sanctions-program jurisdiction "
    "nexus (Iran/North Korea/Syria/Cuba, via country or passport-issuance country) is guaranteed a "
    "Top-5 floor in the final ranking. This reflects the strict-liability nature of OFAC compliance "
    "obligations, which are categorically different from a behavioral-pattern score. Applied to 1 "
    "customer in this dataset (C88888).",
    "",
    "Cluster-diversity cap: 12 customers were found to be members of a single coordinated device+IP "
    "mule ring (Typologies 3 & 9). Left unadjusted, that one incident would occupy 9 of 10 Top-10 slots. "
    "No more than 3 members of any single device/IP-linked cluster are shown as individual Top-10 rows; "
    "all 12 members are fully documented (every txn_id) in the T03/T09 sheets and must be escalated as "
    "one linked case regardless of this cap.",
    "",
    "Calibration notes carried from the case brief:",
    "  - Typology 8 (cash-in/cash-out): 489 of 501 customers (97.6%) have SOME txn in both MCC 6011 "
    "(ATM) and MCC 4829 (money transfer) -- this raw pair is NOT discriminating and is not used for "
    "ranking. Only pairs within <=72h with a remit/ATM ratio >=80% are scored; this reduced the "
    "qualifying population to 1 customer / 1 pair.",
    "  - Typology 4 (geo-hopping): no customer in this dataset is literally 100% exclusive to non-US "
    "high-risk-MCC activity (max observed: 48%); ranking uses cross-border share (continuous) plus "
    "raw volume/amount instead of a hard exclusivity filter.",
    "  - Typology 10 (self-merchant): 175 of 194 UBO records share a full-name match with some customer "
    "in this dataset, but 0 of those are corroborated by a second identity attribute (date of birth). "
    "This is attributable to the dataset's limited name-generation pool (customers and UBOs drawn from "
    "the same first/last-name lists), not real identity links. Self-merchant behavior is NOT CONFIRMED "
    "in this dataset; raw name-only collisions are reported in the T10 sheet for audit transparency "
    "only, explicitly marked unconfirmed, and contribute no hits to the customer ranking.",
    "  - Typology 6 (ECOM w/o 3DS) and Typology 7 (chargeback outlier): evaluated over the FULL dataset "
    "period rather than Task 3.1's rolling 30-day monitoring window, since this is a retrospective "
    "investigation rather than a live daily feed.",
]
for line in methodology_lines:
    ws.cell(row=row, column=1, value=line).font = BODY_FONT
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    row += 1
row += 1
ws.cell(row=row, column=1, value="Typology Confirmation Summary").font = SUBTITLE_FONT
row += 1
confirm_rows = [
    ("T01 Structuring", "104 txns $980-995 in high-risk MCCs (brief-confirmed) -> 6 customers cluster "
                         "at >=3 txns/7 days"),
    ("T02 Card Testing", "254 sub-$5 ECOM txns, 70% decline, 22 unique cards (brief-confirmed) -> "
                          "8 qualifying 5-attempt/30-min/3-merchant windows with an approval"),
    ("T03 Device Sharing", "12-customer extreme case on device dev_shared_4829 CONFIRMED (66-min full "
                            "session; strict 60-min window alone caps at 11)"),
    ("T04 Geo-Hopping X-Border", "7,479 cross-border txns / 23.1% CONFIRMED; 342 customers with >=3 "
                                  "high-risk-MCC cross-border txns (no customer literally 100% exclusive)"),
    ("T05 PEP + High-Risk MCC", "16 PEPs CONFIRMED, all transact at high-risk MCC merchants; 341 txns "
                                 "exceed 2x expected ticket"),
    ("T06 ECOM w/o 3DS", "5,983 KEYED txns / 33.8% of ECOM CONFIRMED; 297 customers with >=5 over the "
                          "full period"),
    ("T07 Chargeback Outlier", "M3030 (ElectroHub #330) CONFIRMED at 12.0% ratio vs. portfolio baseline"),
    ("T08 Cash-In/Cash-Out", "489/501 (97.6%) raw MCC-pair CONFIRMED as context-only, non-discriminating; "
                              "1 customer / 1 pair qualifies under tightened <=72h & >=80%-ratio criteria"),
    ("T09 IP Ring", "12-customer extreme case on IP 45.129.55.210 CONFIRMED"),
    ("T10 Self-Merchant", "NOT CONFIRMED -- 175 raw name collisions, 0 DOB-corroborated (see calibration "
                           "note above)"),
    ("T11 FATF/OFAC Jurisdiction", "3 customers CONFIRMED (1 Iranian passport [sanctions_match_score "
                                    "0.84], 2 Venezuela); 42 cross-border txns"),
]
for label, val in confirm_rows:
    ws.cell(row=row, column=1, value=f"{label}: {val}").font = BODY_FONT
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    row += 1
ws.row_dimensions.default_row_height = 15
for rr in range(1, row):
    ws.row_dimensions[rr].height = 30 if rr > 3 else 20

# --- Top 10 Customers sheet -------------------------------------------------
top10_export = TOP10_FULL[[
    "customer_id", "full_name", "final_score", "n_distinct_typologies", "typologies",
    "total_amount", "txn_count", "period_start", "period_end", "country", "risk_rating",
    "pep_flag", "kyc_level", "regulatory_escalation_floor_applied", "narrative", "referenced_txn_ids",
]].rename(columns={
    "final_score": "suspicion_score", "typologies": "typologies_detected",
    "total_amount": "total_amount_usd_lifetime", "txn_count": "txn_count_lifetime",
    "period_start": "activity_period_start", "period_end": "activity_period_end",
})
top10_export.insert(0, "rank", range(1, len(top10_export) + 1))
write_df_sheet(
    wb, "Top10_Customers", top10_export,
    title="TOP 10 MOST SUSPICIOUS CUSTOMERS",
    currency_cols={"total_amount_usd_lifetime"},
    wrap_cols={"narrative", "referenced_txn_ids", "typologies_detected"},
    col_widths={"rank": 6, "customer_id": 12, "full_name": 18, "suspicion_score": 14,
                "n_distinct_typologies": 10, "typologies_detected": 40, "total_amount_usd_lifetime": 16,
                "txn_count_lifetime": 12, "activity_period_start": 14, "activity_period_end": 14,
                "country": 8, "risk_rating": 10, "pep_flag": 8, "kyc_level": 10,
                "regulatory_escalation_floor_applied": 12, "narrative": 100, "referenced_txn_ids": 60},
)

# --- Top 30 Transactions sheet ----------------------------------------------
top30_export = TOP30[[
    "txn_id", "customer_id", "merchant_id", "dba_name", "mcc", "mcc_description", "amount_usd",
    "txn_timestamp_utc", "merchant_country", "channel", "status", "final_score", "n_typologies",
    "typologies", "notes",
]].rename(columns={"final_score": "suspicion_score", "typologies": "typologies_detected"})
top30_export.insert(0, "rank", range(1, len(top30_export) + 1))
write_df_sheet(
    wb, "Top30_Transactions", top30_export,
    title="TOP 30 MOST SUSPICIOUS TRANSACTIONS",
    currency_cols={"amount_usd"},
    wrap_cols={"notes", "typologies_detected"},
    col_widths={"rank": 6, "txn_id": 10, "customer_id": 12, "merchant_id": 10, "dba_name": 20,
                "mcc": 6, "mcc_description": 22, "amount_usd": 12, "txn_timestamp_utc": 20,
                "merchant_country": 10, "channel": 12, "status": 10, "suspicion_score": 14,
                "n_typologies": 10, "typologies_detected": 45, "notes": 55},
)

print(f"  Wrote Summary, Top10_Customers, Top30_Transactions")

# --- Per-typology sheets -----------------------------------------------------
CURRENCY_HINTS = {"amount_usd", "total_amount", "total_amount_usd", "total_amount_in_window",
                  "total_amount_30d", "atm_amount", "remit_amount", "customer_chargeback_amount",
                  "total_amount_at_own_merchant", "expected_avg_ticket_usd", "hr_xb_amount",
                  "expected_monthly_volume_usd", "total_amount_7d"}
WRAP_HINTS = {"evidence", "txn_ids", "customer_ids", "merchant_ids", "mccs_in_window", "countries",
              "typologies", "notes"}

typology_titles = {
    "T01_Structuring": "TYPOLOGY 1 — STRUCTURING (repeated sub-CTR amounts in high-risk MCC band)",
    "T02_Card_Testing": "TYPOLOGY 2 — CARD TESTING (rapid low-value ECOM probing)",
    "T03_Device_Sharing": "TYPOLOGY 3 — DEVICE SHARING / MULE RING",
    "T04_Geo_Hopping_XBorder": "TYPOLOGY 4 — GEO-HOPPING CROSS-BORDER",
    "T05_PEP_High_Risk_MCC": "TYPOLOGY 5 — HIGH-RISK MCC + PEP COMBINATION",
    "T06_ECOM_No_3DS": "TYPOLOGY 6 — ECOM WITHOUT 3DS IN HIGH-RISK MCC",
    "T07_Chargeback_Outlier": "TYPOLOGY 7 — CHARGEBACK RATIO OUTLIER (merchant-level)",
    "T07_Chargeback_Customer_Link": "TYPOLOGY 7b — CUSTOMERS LINKED TO CHARGEBACK-OUTLIER MERCHANT(S)",
    "T08_Cash_In_Cash_Out": "TYPOLOGY 8 — CASH-IN TO CASH-OUT (tightened: <=72h, >=80% ratio)",
    "T09_IP_Ring": "TYPOLOGY 9 — IP ADDRESS RING",
    "T10_Self_Merchant": "TYPOLOGY 10 — SELF-MERCHANT (name+DOB audit; 0 confirmed, see Summary)",
    "T11_FATF_OFAC_Jurisdiction": "TYPOLOGY 11 — FATF/OFAC JURISDICTION",
}

for key, df in TYPOLOGY_TABLES.items():
    if key == "T07_Chargeback_Customer_Link":
        continue  # combined into the T07 sheet below
    sheet_name = key
    df_out = df.copy()
    currency_cols = {c for c in df_out.columns if c in CURRENCY_HINTS}
    wrap_cols = {c for c in df_out.columns if c in WRAP_HINTS}
    if key == "T07_Chargeback_Outlier":
        # Combine merchant-level outlier summary + customer link table in one sheet
        ws, next_row = write_df_sheet(
            wb, sheet_name, df_out, title=typology_titles.get(key, key),
            currency_cols=currency_cols, wrap_cols=wrap_cols,
        )
        link_df = TYPOLOGY_TABLES.get("T07_Chargeback_Customer_Link")
        if link_df is not None and len(link_df):
            link_currency = {c for c in link_df.columns if c in CURRENCY_HINTS}
            link_wrap = {c for c in link_df.columns if c in WRAP_HINTS}
            write_table_on_ws(
                ws, link_df, title=typology_titles.get("T07_Chargeback_Customer_Link"),
                currency_cols=link_currency, wrap_cols=link_wrap, start_row=next_row + 2,
            )
        print(f"  Wrote {sheet_name} ({len(df_out)} merchant rows + "
              f"{len(link_df) if link_df is not None else 0} customer-link rows)")
        continue
    write_df_sheet(
        wb, sheet_name, df_out, title=typology_titles.get(key, key),
        currency_cols=currency_cols, wrap_cols=wrap_cols,
    )
    print(f"  Wrote {sheet_name} ({len(df_out)} rows)")

# --- Full ranked customer table (audit trail beyond the curated Top 10) ----
full_rank_export = cust_score[[
    "customer_id", "full_name", "final_score", "n_distinct_typologies", "typologies",
    "total_amount", "txn_count", "country", "risk_rating", "pep_flag",
    "regulatory_escalation_floor_applied",
]].rename(columns={"final_score": "suspicion_score", "typologies": "typologies_detected",
                    "total_amount": "total_amount_usd_lifetime"})
full_rank_export.insert(0, "rank", range(1, len(full_rank_export) + 1))
write_df_sheet(
    wb, "Full_Ranked_Customers", full_rank_export,
    title="FULL RANKED CUSTOMER LIST (all customers with >=1 typology hit; audit trail behind "
          "the curated Top 10)",
    currency_cols={"total_amount_usd_lifetime"}, wrap_cols={"typologies_detected"},
    col_widths={"rank": 6, "customer_id": 12, "full_name": 18, "suspicion_score": 14,
                "n_distinct_typologies": 10, "typologies_detected": 45,
                "total_amount_usd_lifetime": 16, "txn_count": 10},
)
print(f"  Wrote Full_Ranked_Customers ({len(full_rank_export)} rows)")

# Reorder: Summary first, then Top10/Top30, then typologies, then audit trail
sheet_order = ["Summary", "Top10_Customers", "Top30_Transactions"] + [
    k for k in TYPOLOGY_TABLES.keys() if k != "T07_Chargeback_Customer_Link"
] + ["Full_Ranked_Customers"]
wb._sheets = [wb[s] for s in sheet_order if s in wb.sheetnames]

import os
os.makedirs("/mnt/user-data/outputs", exist_ok=True)
wb.save(OUTPUT_XLSX)
print(f"\nSaved workbook: {OUTPUT_XLSX}")
print(f"Sheets: {wb.sheetnames}")
