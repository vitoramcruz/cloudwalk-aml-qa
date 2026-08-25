"""
CloudWalk Payments Inc. — AML Case File
SAR Evidence Preparation: Transaction Timeline, Relationship Network,
and CDD-Deviation Analysis

For the top 3 most suspicious customers identified in the prior composite-
score ranking (Task 4.1): C12105, C12451, C12373 -- the three highest-scoring
individual representatives of the confirmed 12-account device/IP mule ring.

Produces, per customer:
  1. Full chronological transaction timeline with inline suspicious-activity
     flags (device sharing, IP ring, geo-hopping, ECOM-without-3DS,
     structuring-band, PEP+high-risk-MCC, chargeback, cash-in/cash-out,
     FATF/OFAC jurisdiction).
  2. Relationship network: other customers sharing the same device_id(s) or
     ip_address(es), and merchants where multiple flagged customers transact.
  3. Transactional profile vs. CDD: actual monthly volume and average ticket
     vs. customers_kyc.expected_monthly_volume_usd / expected_avg_ticket_usd,
     with percentage deviation and a plain-language interpretation.

All output in English (SAR-evidence use). Prepared by: AML/BSA Compliance — Investigations.
"""

import json
import warnings
from datetime import timedelta

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

# ============================================================================
# CONFIG
# ============================================================================
INPUT_FILE = "/mnt/project/AMLFT_Analyst_JIM__1_.xlsx"
TARGET_CUSTOMERS = ["C12105", "C12451", "C12373"]  # rank 1, 2, 3 from Task 4.1
HIGH_RISK_MCCS = [4829, 5944, 5967, 6011, 6051, 7995]
OFAC_BLOCK = ["IR", "KP", "SY", "CU"]
FATF_GREY = ["VE"]
STRUCT_LOW, STRUCT_HIGH = 980.00, 995.00
PEP_MULTIPLIER = 2.0
CB_ALERT_THRESHOLD = 0.05
CIO_WINDOW = timedelta(hours=72)
CIO_RATIO = 0.80

# ============================================================================
# LOAD DATA
# ============================================================================
print("Loading workbook...")
xls = pd.ExcelFile(INPUT_FILE)
tx = pd.read_excel(xls, "Transactions")
cust = pd.read_excel(xls, "Customers_KYC")
merch = pd.read_excel(xls, "Merchants_KYB")
cards = pd.read_excel(xls, "Cards")

tx["txn_timestamp_utc"] = pd.to_datetime(tx["txn_timestamp_utc"])
DATASET_MAX_TS = tx["txn_timestamp_utc"].max()
DATASET_MIN_TS = tx["txn_timestamp_utc"].min()

merch_small = merch[["merchant_id", "dba_name", "primary_mcc", "high_risk_mcc_flag"]]
cust_small = cust[[
    "customer_id", "full_name", "country", "doc_issue_country", "pep_flag",
    "sanctions_match_score", "risk_rating", "expected_avg_ticket_usd",
    "expected_monthly_volume_usd", "kyc_level",
]]
txe = (tx.merge(merch_small, on="merchant_id", how="left", validate="m:1")
         .merge(cust_small, on="customer_id", how="left", validate="m:1"))

print(f"  {len(tx):,} txns | {len(cust):,} customers | {len(merch):,} merchants | {len(cards):,} cards")
print(f"  Period: {DATASET_MIN_TS.date()} to {DATASET_MAX_TS.date()}")

# ============================================================================
# PRE-COMPUTE DATASET-WIDE SIGNALS NEEDED FOR FLAGGING
# ============================================================================
# Device sharing: TRUE mule-ring signal requires >=3 DISTINCT customers on the
# SAME device within a tight time window (Query 2 params: 60 min), not merely
# >=3 different customers having ever used that device_id across the full
# 8-month period -- the latter is common (669 devices) and not discriminating,
# the same false-positive trap identified in Task 4.1. Re-derive the genuinely
# time-clustered device-sharing incidents here rather than a raw lifetime count.
device_cust_hr = txe[txe["mcc"].isin(HIGH_RISK_MCCS) & txe["device_id"].notna()]
DS_WINDOW = timedelta(minutes=60)
DS_MIN_CUST = 3

def find_device_sharing_incidents(df):
    """Sliding forward-window scan per device_id; returns list of
    (device_id, txn_ids_in_incident, customer_ids_in_incident) for every
    window where >=DS_MIN_CUST distinct customers transact within DS_WINDOW."""
    incidents = []
    for dev, gdf in df.groupby("device_id"):
        gdf = gdf.sort_values("txn_timestamp_utc").reset_index(drop=True)
        ts = gdf["txn_timestamp_utc"].values
        n = len(gdf)
        for i in range(n):
            end_ts = ts[i] + np.timedelta64(int(DS_WINDOW.total_seconds()), "s")
            j = i
            while j < n and ts[j] <= end_ts:
                j += 1
            wdf = gdf.iloc[i:j]
            if wdf["customer_id"].nunique() >= DS_MIN_CUST:
                incidents.append((dev, set(wdf["txn_id"]), set(wdf["customer_id"])))
    return incidents

device_incidents = find_device_sharing_incidents(device_cust_hr)
# Also reconcile the confirmed extreme case at its full observed extent (see
# Task 4.1: the 12-customer dev_shared_4829 session spans 66 min, one minute
# past the strict 60-min window, and is reported at full extent there).
mcc4829_dev_counts = txe[txe["mcc"] == 4829].groupby("device_id")["customer_id"].nunique()
for dev, raw_n in mcc4829_dev_counts[mcc4829_dev_counts >= 5].items():
    windowed_n = max((len(c) for d, t, c in device_incidents if d == dev), default=0)
    if raw_n > windowed_n:
        full = txe[(txe["device_id"] == dev) & (txe["mcc"] == 4829)]
        device_incidents.append((dev, set(full["txn_id"]), set(full["customer_id"])))

# txn_id -> device incident membership (device_id, full customer set)
txnid_to_device_incident = {}
for dev, txn_ids, cust_ids in device_incidents:
    for t in txn_ids:
        # keep the largest incident if a txn appears in more than one window
        if t not in txnid_to_device_incident or len(cust_ids) > len(txnid_to_device_incident[t][1]):
            txnid_to_device_incident[t] = (dev, cust_ids)

print(f"  Time-clustered device-sharing incidents found (>=3 customers/60min, "
      f"high-risk MCC): {len(device_incidents)}")
print(f"  Distinct transactions implicated: {len(txnid_to_device_incident)}")

# IP ring: distinct customers per ip_address in ECOM
ecom_ip = txe[(txe["channel"] == "ECOM") & txe["ip_address"].notna()]
ip_cust_counts = ecom_ip.groupby("ip_address")["customer_id"].nunique()
shared_ips = set(ip_cust_counts[ip_cust_counts >= 3].index)  # Query 8 threshold

# Chargeback outlier merchant(s)
merch_ratios = txe.groupby("merchant_id").agg(total=("txn_id", "count"), cb=("is_chargeback", "sum"))
merch_ratios["ratio"] = merch_ratios["cb"] / merch_ratios["total"]
outlier_merchants = set(merch_ratios[merch_ratios["ratio"] > CB_ALERT_THRESHOLD].index)

print(f"  Confirmed shared IPs (>=3 customers, ECOM): {len(shared_ips)}")
print(f"  Chargeback-outlier merchants (>5% ratio): {sorted(outlier_merchants)}")


def flag_row(r):
    """Return a list of typology flags applicable to a single transaction row."""
    flags = []
    if r["txn_id"] in txnid_to_device_incident:
        flags.append("DEVICE_SHARING")
    if pd.notna(r["ip_address"]) and r["ip_address"] in shared_ips and r["channel"] == "ECOM":
        flags.append("IP_RING")
    if r["mcc"] in HIGH_RISK_MCCS and r["merchant_country"] != "US":
        flags.append("GEO_HOPPING_XBORDER")
    if r["channel"] == "ECOM" and r["pos_entry_mode"] == "KEYED" and r["high_risk_mcc_flag"] is True:
        flags.append("ECOM_NO_3DS")
    if STRUCT_LOW <= r["amount_usd"] <= STRUCT_HIGH and r["mcc"] in HIGH_RISK_MCCS:
        flags.append("STRUCTURING_BAND")
    if r["channel"] == "ECOM" and r["amount_usd"] < 5.00:
        flags.append("CARD_TESTING_LOW_VALUE")
    if bool(r["pep_flag"]) and r["high_risk_mcc_flag"] is True and pd.notna(r["expected_avg_ticket_usd"]) \
            and r["expected_avg_ticket_usd"] > 0 and r["amount_usd"] > PEP_MULTIPLIER * r["expected_avg_ticket_usd"]:
        flags.append("PEP_ABOVE_EXPECTED_TICKET")
    if bool(r["is_chargeback"]) and r["merchant_id"] in outlier_merchants:
        flags.append("CHARGEBACK_AT_OUTLIER_MERCHANT")
    if r["mcc"] in (6011, 4829):
        flags.append("CASH_IN_CASH_OUT_CONTEXT")  # context flag; true pair confirmed separately
    if r["merchant_country"] != "US" and (
        r["country"] in OFAC_BLOCK + FATF_GREY or r["doc_issue_country"] in OFAC_BLOCK + FATF_GREY
    ):
        flags.append("FATF_OFAC_JURISDICTION")
    return flags


# ============================================================================
# PER-CUSTOMER: TIMELINE, RELATIONSHIP NETWORK, CDD DEVIATION
# ============================================================================
TIMELINES = {}
RELATIONSHIP_NETWORKS = {}
CDD_ROWS = []
CIO_PAIRS_BY_CUST = {}

for rank, cid in enumerate(TARGET_CUSTOMERS, start=1):
    print(f"\n{'=' * 70}\nCUSTOMER #{rank}: {cid}\n{'=' * 70}")
    cust_row = cust[cust["customer_id"] == cid].iloc[0]
    sub = txe[txe["customer_id"] == cid].sort_values("txn_timestamp_utc").reset_index(drop=True)
    print(f"  {cust_row['full_name']} | country: {cust_row['country']} | "
          f"risk_rating: {cust_row['risk_rating']} | PEP: {cust_row['pep_flag']} | "
          f"KYC level: {cust_row['kyc_level']}")
    print(f"  Total transactions: {len(sub)} | Period: {sub['txn_timestamp_utc'].min()} "
          f"to {sub['txn_timestamp_utc'].max()}")

    # --- 1. Timeline with inline flags -------------------------------------
    sub = sub.copy()
    sub["flags"] = sub.apply(flag_row, axis=1)
    sub["flags_str"] = sub["flags"].apply(lambda fl: "; ".join(f"[FLAG: {f}]" for f in fl) if fl else "")
    n_flagged = (sub["flags"].apply(len) > 0).sum()
    print(f"  Flagged transactions: {n_flagged} of {len(sub)}")

    timeline_cols = [
        "txn_id", "txn_timestamp_utc", "merchant_id", "dba_name", "mcc", "primary_mcc",
        "amount_usd", "status", "channel", "pos_entry_mode", "device_id", "ip_address",
        "merchant_country", "is_chargeback", "flags_str",
    ]
    TIMELINES[cid] = sub[timeline_cols].rename(columns={"flags_str": "flags"})

    # --- Cash-in/cash-out pair check (tightened <=72h, >=80% ratio) --------
    atm = sub[(sub["mcc"] == 6011) & (sub["status"] == "approved") & (sub["amount_usd"] >= 500)]
    remit = sub[(sub["mcc"] == 4829) & (sub["status"] == "approved") & (sub["amount_usd"] >= 500)]
    cio_pairs = []
    for _, a in atm.iterrows():
        window_r = remit[(remit["txn_timestamp_utc"] > a["txn_timestamp_utc"])
                         & (remit["txn_timestamp_utc"] <= a["txn_timestamp_utc"] + CIO_WINDOW)]
        for _, rr in window_r.iterrows():
            ratio = rr["amount_usd"] / a["amount_usd"]
            if ratio >= CIO_RATIO:
                cio_pairs.append((a["txn_id"], rr["txn_id"], ratio,
                                  (rr["txn_timestamp_utc"] - a["txn_timestamp_utc"]).total_seconds() / 3600))
    CIO_PAIRS_BY_CUST[cid] = cio_pairs
    if cio_pairs:
        print(f"  Cash-in/cash-out qualifying pairs (<=72h, >=80% ratio): {len(cio_pairs)}")

    # --- 2. Relationship network --------------------------------------------
    own_devices = sorted(set(d for d in sub["device_id"].dropna().unique()))
    own_ips = sorted(set(i for i in sub["ip_address"].dropna().unique()))

    device_links = []
    linked_via_device = {}
    for t in sub.loc[sub["txn_id"].isin(txnid_to_device_incident.keys()), "txn_id"]:
        dev, cust_ids = txnid_to_device_incident[t]
        for other_cid in cust_ids:
            if other_cid == cid:
                continue
            linked_via_device.setdefault((dev, other_cid), 0)
            linked_via_device[(dev, other_cid)] += 1
    for (dev, other_cid), cnt in linked_via_device.items():
        name_match = cust.loc[cust.customer_id == other_cid, "full_name"]
        device_links.append({
            "device_id": dev, "linked_customer_id": other_cid,
            "linked_customer_name": name_match.iloc[0] if len(name_match) else None,
            "shared_incident_txn_overlap": cnt,
        })

    ip_links = []
    for ip in own_ips:
        if ip in shared_ips:
            co_users = ecom_ip[(ecom_ip["ip_address"] == ip) & (ecom_ip["customer_id"] != cid)]
            for other_cid, g in co_users.groupby("customer_id"):
                ip_links.append({
                    "ip_address": ip, "linked_customer_id": other_cid,
                    "linked_customer_name": cust.loc[cust.customer_id == other_cid, "full_name"].iloc[0]
                        if (cust.customer_id == other_cid).any() else None,
                    "shared_txn_count": len(g),
                    "first_seen": g["txn_timestamp_utc"].min(), "last_seen": g["txn_timestamp_utc"].max(),
                })

    device_link_df = pd.DataFrame(device_links).sort_values(
        "shared_incident_txn_overlap", ascending=False) if device_links else pd.DataFrame(
        columns=["device_id", "linked_customer_id", "linked_customer_name", "shared_incident_txn_overlap"])
    ip_link_df = pd.DataFrame(ip_links).sort_values(
        "shared_txn_count", ascending=False) if ip_links else pd.DataFrame(
        columns=["ip_address", "linked_customer_id", "linked_customer_name", "shared_txn_count",
                 "first_seen", "last_seen"])

    n_linked_via_device = device_link_df["linked_customer_id"].nunique() if len(device_link_df) else 0
    n_linked_via_ip = ip_link_df["linked_customer_id"].nunique() if len(ip_link_df) else 0
    print(f"  Distinct customers linked via shared device: {n_linked_via_device}")
    print(f"  Distinct customers linked via shared IP: {n_linked_via_ip}")

    # Merchants where multiple TARGET customers (this analysis's 3 suspects) transact
    # -- restricted to HIGH-RISK MCC merchants. With only 100 merchants total in
    # this portfolio, ordinary merchant overlap (grocery, retail, etc.) across any
    # two customers is expected by chance and not a meaningful AML signal; overlap
    # specifically at high-risk-MCC (money-transfer/quasi-cash/ATM) merchants is
    # the operationally relevant convergence point.
    own_merchants_hr = set(sub.loc[sub["high_risk_mcc_flag"] == True, "merchant_id"].unique())
    own_merchants_all = set(sub["merchant_id"].unique())
    shared_target_merchants = []
    for other in TARGET_CUSTOMERS:
        if other == cid:
            continue
        other_sub = txe[txe["customer_id"] == other]
        other_merchants_hr = set(other_sub.loc[other_sub["high_risk_mcc_flag"] == True, "merchant_id"].unique())
        other_merchants_all = set(other_sub["merchant_id"].unique())
        common_hr = own_merchants_hr & other_merchants_hr
        common_all = own_merchants_all & other_merchants_all
        for m in common_all:
            shared_target_merchants.append({
                "merchant_id": m, "dba_name": merch.loc[merch.merchant_id == m, "dba_name"].iloc[0],
                "high_risk_mcc_merchant": m in common_hr,
                "co_transacting_customer_id": other,
            })
    shared_merch_df = pd.DataFrame(shared_target_merchants).drop_duplicates() if shared_target_merchants \
        else pd.DataFrame(columns=["merchant_id", "dba_name", "high_risk_mcc_merchant", "co_transacting_customer_id"])
    n_hr_shared = shared_merch_df["high_risk_mcc_merchant"].sum() if len(shared_merch_df) else 0
    print(f"  Merchants shared with the other 2 target customers: "
          f"{shared_merch_df['merchant_id'].nunique() if len(shared_merch_df) else 0} total "
          f"({n_hr_shared} at high-risk MCC merchants -- the operationally relevant overlap)")

    RELATIONSHIP_NETWORKS[cid] = {
        "device_links": device_link_df, "ip_links": ip_link_df, "shared_merchants": shared_merch_df,
        "own_devices": own_devices, "own_ips": own_ips,
    }

    # --- 3. Transactional profile vs. CDD -----------------------------------
    period_days = (sub["txn_timestamp_utc"].max() - sub["txn_timestamp_utc"].min()).days
    period_months = max(period_days / 30.44, 1e-6)
    total_amount = sub["amount_usd"].sum()
    txn_count = len(sub)
    actual_monthly_volume = total_amount / period_months
    actual_avg_ticket = total_amount / txn_count

    expected_monthly_volume = cust_row["expected_monthly_volume_usd"]
    expected_avg_ticket = cust_row["expected_avg_ticket_usd"]

    vol_deviation_pct = (actual_monthly_volume - expected_monthly_volume) / expected_monthly_volume * 100 \
        if expected_monthly_volume else np.nan
    ticket_deviation_pct = (actual_avg_ticket - expected_avg_ticket) / expected_avg_ticket * 100 \
        if expected_avg_ticket else np.nan

    def interpret(pct, label):
        if pd.isna(pct):
            return f"{label}: expected value not available for comparison."
        if pct >= 100:
            return f"{label} is {pct:.0f}% ABOVE the CDD-declared expectation (more than double) -- material deviation warranting profile review."
        if pct >= 25:
            return f"{label} is {pct:.0f}% above the CDD-declared expectation -- moderate deviation."
        if pct <= -50:
            return f"{label} is {abs(pct):.0f}% BELOW the CDD-declared expectation -- account may be materially under-utilized relative to stated profile, or profile was overstated at onboarding."
        return f"{label} is within {abs(pct):.0f}% of the CDD-declared expectation -- no material deviation on this metric alone."

    cdd_row = {
        "customer_id": cid, "full_name": cust_row["full_name"],
        "kyc_level": cust_row["kyc_level"], "risk_rating": cust_row["risk_rating"],
        "period_days_observed": period_days, "txn_count": txn_count, "total_amount_usd": total_amount,
        "actual_monthly_volume_usd": actual_monthly_volume, "expected_monthly_volume_usd": expected_monthly_volume,
        "monthly_volume_deviation_pct": vol_deviation_pct,
        "actual_avg_ticket_usd": actual_avg_ticket, "expected_avg_ticket_usd": expected_avg_ticket,
        "avg_ticket_deviation_pct": ticket_deviation_pct,
        "volume_interpretation": interpret(vol_deviation_pct, "Actual monthly volume"),
        "ticket_interpretation": interpret(ticket_deviation_pct, "Actual average ticket"),
    }
    CDD_ROWS.append(cdd_row)
    print(f"  Actual monthly volume: ${actual_monthly_volume:,.2f} vs expected "
          f"${expected_monthly_volume:,.2f} ({vol_deviation_pct:+.0f}%)")
    print(f"  Actual avg ticket: ${actual_avg_ticket:,.2f} vs expected ${expected_avg_ticket:,.2f} "
          f"({ticket_deviation_pct:+.0f}%)")

CDD_DF = pd.DataFrame(CDD_ROWS)

# ============================================================================
# EXCEL EXPORT
# ============================================================================
print("\n" + "=" * 70)
print("EXPORTING EXCEL WORKBOOK")
print("=" * 70)

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

os.makedirs("/mnt/user-data/outputs", exist_ok=True)
OUTPUT_XLSX = "/mnt/user-data/outputs/SAR_Evidence_Package_Top3.xlsx"

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="7A1F1F", end_color="7A1F1F", fill_type="solid")
FLAG_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="7A1F1F")
SUBTITLE_FONT = Font(name="Arial", bold=True, size=11, color="7A1F1F")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def clean_val(val):
    if isinstance(val, (list, set)):
        return ", ".join(str(x) for x in val)
    if isinstance(val, pd.Timestamp):
        return val.tz_localize(None) if val.tzinfo else val
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        return float(val) if not np.isnan(val) else None
    if isinstance(val, (np.bool_,)):
        return bool(val)
    return val


def write_table_on_ws(ws, df, title=None, currency_cols=None, wrap_cols=None,
                       col_widths=None, start_row=1, highlight_flag_col=None):
    r = start_row
    if title:
        ws.cell(row=r, column=1, value=title).font = TITLE_FONT
        r += 2
    if df is None or len(df) == 0:
        ws.cell(row=r, column=1, value="(No records.)").font = BODY_FONT
        return r + 1
    header_row = r
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=header_row, column=j, value=str(col))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    last_i = header_row
    for i, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        last_i = i
        is_flagged = highlight_flag_col and bool(row.get(highlight_flag_col, "")) and str(row.get(highlight_flag_col, "")).strip() != ""
        for j, col in enumerate(df.columns, start=1):
            val = clean_val(row[col])
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            if currency_cols and col in currency_cols:
                c.number_format = '$#,##0.00'
            if wrap_cols and col in wrap_cols:
                c.alignment = WRAP_ALIGN
            if is_flagged:
                c.fill = FLAG_FILL
    if start_row == 1 or ws.freeze_panes is None:
        try:
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
        except Exception:
            pass
    widths = col_widths or {}
    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        new_w = widths.get(col, 16)
        cur_w = ws.column_dimensions[letter].width
        if cur_w is None or new_w > cur_w:
            ws.column_dimensions[letter].width = new_w
    return last_i + 1


def write_df_sheet(wb, sheet_name, df, **kwargs):
    ws = wb.create_sheet(sheet_name[:31])
    next_row = write_table_on_ws(ws, df, **kwargs)
    return ws, next_row


wb = openpyxl.Workbook()
wb.remove(wb.active)

# --- Summary sheet -----------------------------------------------------------
ws = wb.create_sheet("Summary")
ws.column_dimensions["A"].width = 105
row = 1
ws.cell(row=row, column=1, value="CloudWalk Payments Inc. — SAR Evidence Package (Top 3 Suspects)").font = TITLE_FONT
row += 1
ws.cell(row=row, column=1, value="Prepared for federal SAR filing. Source: Task 4.1 composite-score "
                                  "ranking. Targets: C12105 (rank #1), C12451 (rank #2), C12373 (rank #3), "
                                  "the three highest-scoring individual members of the confirmed 12-account "
                                  "device/IP mule ring.").font = BODY_FONT
row += 2
ws.cell(row=row, column=1, value="Key Evidentiary Finding").font = SUBTITLE_FONT
row += 1
key_finding = (
    "All three targets, plus 9 additional linked accounts (12 total), transacted on 2025-10-26 between "
    "13:59:59 and 15:05:59 UTC (66-minute span, exactly 6-minute cadence) using the SAME device_id "
    "(dev_shared_4829), the SAME IP address (45.129.55.210), at the SAME merchant (M3061, 'RemitFast #361', "
    "MCC 4829 - Money Transfer). This is a single, scripted, coordinated session consistent with a mule-"
    "account network, not independent customer behavior."
)
ws.cell(row=row, column=1, value=key_finding).font = Font(name="Arial", size=10, bold=True)
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
row += 3
ws.cell(row=row, column=1, value="Per-Customer Headline Metrics").font = SUBTITLE_FONT
row += 1
for _, r in CDD_DF.iterrows():
    line = (f"{r['customer_id']} ({r['full_name']}): {r['txn_count']} txns observed; "
            f"avg ticket ${r['actual_avg_ticket_usd']:,.2f} vs CDD-expected ${r['expected_avg_ticket_usd']:,.2f} "
            f"({r['avg_ticket_deviation_pct']:+.0f}%); monthly volume ${r['actual_monthly_volume_usd']:,.2f} "
            f"vs CDD-expected ${r['expected_monthly_volume_usd']:,.2f} ({r['monthly_volume_deviation_pct']:+.0f}%)")
    ws.cell(row=row, column=1, value=line).font = BODY_FONT
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    row += 1
row += 1
ws.cell(row=row, column=1, value="Methodology Notes").font = SUBTITLE_FONT
row += 1
notes = [
    "Device-sharing and IP-ring flags use the SAME time-windowed logic validated in Task 4.1 "
    "(>=3 distinct customers within a 60-minute forward window at a high-risk MCC for device; >=3 "
    "distinct customers on one IP in ECOM). A raw lifetime device-reuse count (ignoring time) is NOT "
    "used, since it flags 669 of the portfolio's devices and is not discriminating.",
    "Merchant-overlap ('relationship network') is reported both as raw overlap and restricted to "
    "high-risk-MCC merchants. With only 41 high-risk-MCC merchants and 100 total merchants in this "
    "portfolio, and each target customer individually active across dozens of merchants over 8 months, "
    "high overlap counts are expected by chance and are NOT independently probative -- the device/IP/"
    "merchant/timestamp co-occurrence documented above IS probative because it is simultaneous.",
    "CDD deviation: actual monthly volume = total observed amount / observed active months; actual "
    "average ticket = total observed amount / transaction count. Both are compared against the "
    "customers_kyc.expected_monthly_volume_usd and expected_avg_ticket_usd fields captured at onboarding.",
]
for n in notes:
    ws.cell(row=row, column=1, value=n).font = BODY_FONT
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    row += 1
for rr in range(1, row):
    ws.row_dimensions[rr].height = 30 if rr > 2 else 20

# --- Per-customer Timeline / Relationship / CDD sheets -----------------------
TIMELINE_CURRENCY = {"amount_usd"}
TIMELINE_WRAP = {"flags", "dba_name"}

for cid in TARGET_CUSTOMERS:
    tdf = TIMELINES[cid].copy()
    ws, _ = write_df_sheet(
        wb, f"Timeline_{cid}", tdf,
        title=f"FULL TRANSACTION TIMELINE — {cid} ({cust.loc[cust.customer_id==cid,'full_name'].iloc[0]})",
        currency_cols=TIMELINE_CURRENCY, wrap_cols=TIMELINE_WRAP,
        col_widths={"txn_id": 10, "txn_timestamp_utc": 20, "merchant_id": 10, "dba_name": 20,
                    "mcc": 6, "primary_mcc": 10, "amount_usd": 11, "status": 10, "channel": 12,
                    "pos_entry_mode": 12, "device_id": 16, "ip_address": 15, "merchant_country": 10,
                    "is_chargeback": 10, "flags": 55},
        highlight_flag_col="flags",
    )
    print(f"  Wrote Timeline_{cid} ({len(tdf)} rows, "
          f"{(tdf['flags'].astype(str).str.len() > 0).sum()} flagged)")

# --- Relationship network sheet (all 3 customers combined) ------------------
ws = wb.create_sheet("Relationship_Network")
r = 1
ws.cell(row=r, column=1, value="RELATIONSHIP NETWORK — DEVICE / IP / MERCHANT LINKAGES").font = TITLE_FONT
r += 2
for cid in TARGET_CUSTOMERS:
    net = RELATIONSHIP_NETWORKS[cid]
    name = cust.loc[cust.customer_id == cid, "full_name"].iloc[0]
    ws.cell(row=r, column=1, value=f"{cid} — {name}").font = SUBTITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value=f"Own device_id(s): {', '.join(net['own_devices'])}").font = BODY_FONT
    r += 1
    ws.cell(row=r, column=1, value=f"Own ip_address(es): {', '.join(net['own_ips'])}").font = BODY_FONT
    r += 2
    r = write_table_on_ws(ws, net["device_links"], title="Linked customers (shared device, same time-clustered incident)",
                           col_widths={"device_id": 16, "linked_customer_id": 14, "linked_customer_name": 18,
                                       "shared_incident_txn_overlap": 12}, start_row=r)
    r += 1
    r = write_table_on_ws(ws, net["ip_links"], title="Linked customers (shared IP, ECOM)",
                           col_widths={"ip_address": 16, "linked_customer_id": 14, "linked_customer_name": 18,
                                       "shared_txn_count": 12}, start_row=r,
                           currency_cols=None, wrap_cols=None)
    r += 1
    r = write_table_on_ws(ws, net["shared_merchants"].head(15),
                           title="Merchants shared with other 2 targets (top 15; high_risk_mcc_merchant flag shown; "
                                 "see Summary sheet for the non-discriminating-overlap caveat)",
                           col_widths={"merchant_id": 10, "dba_name": 22, "high_risk_mcc_merchant": 14,
                                       "co_transacting_customer_id": 16}, start_row=r)
    r += 2

# --- CDD Profile Analysis sheet ----------------------------------------------
cdd_export = CDD_DF.copy()
write_df_sheet(
    wb, "CDD_Profile_Analysis", cdd_export,
    title="TRANSACTIONAL PROFILE vs. CDD — VOLUME & TICKET DEVIATION",
    currency_cols={"total_amount_usd", "actual_monthly_volume_usd", "expected_monthly_volume_usd",
                   "actual_avg_ticket_usd", "expected_avg_ticket_usd"},
    wrap_cols={"volume_interpretation", "ticket_interpretation"},
    col_widths={"customer_id": 12, "full_name": 18, "kyc_level": 10, "risk_rating": 10,
                "period_days_observed": 12, "txn_count": 10, "total_amount_usd": 14,
                "actual_monthly_volume_usd": 16, "expected_monthly_volume_usd": 16,
                "monthly_volume_deviation_pct": 14, "actual_avg_ticket_usd": 14,
                "expected_avg_ticket_usd": 14, "avg_ticket_deviation_pct": 12,
                "volume_interpretation": 55, "ticket_interpretation": 55},
)

sheet_order = ["Summary"] + [f"Timeline_{c}" for c in TARGET_CUSTOMERS] + \
    ["Relationship_Network", "CDD_Profile_Analysis"]
wb._sheets = [wb[s] for s in sheet_order if s in wb.sheetnames]

wb.save(OUTPUT_XLSX)
print(f"\nSaved workbook: {OUTPUT_XLSX}")
print(f"Sheets: {wb.sheetnames}")
